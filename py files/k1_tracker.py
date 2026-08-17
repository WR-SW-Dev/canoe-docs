#!/usr/bin/env python3
"""
k1_tracker.py -- Track which K-1s have arrived, per fund / entity / tax year.

The tax-document sibling of statement_tracker.py, and deliberately built to the same
shape: metadata-and-rules only (it reads Canoe's structured document metadata via
GET /v1/documents/data and never opens a document body), an auto-seeded editable
schedule, a dated workbook for the team plus a backend/ folder of supporting detail,
and the same two destinations (--graph for the unattended job, --dest for a local
archive). Anything genuinely generic -- the metadata endpoint's retry/rate-limit
quirks, archive hyperlink resolution, the local-archive walk -- is imported from
statement_tracker rather than copied, so a fix there fixes both trackers.

What is different, because K-1s are not statements
-------------------------------------------------
* No cadence to infer. Every K-1 is annual (445 of 448 carry a 12-31 data date), so
  the statement tracker's frequency machinery is replaced by a single question:
  which fund x entity x tax year combinations SHOULD have a K-1?

* Expectation comes from HOLDINGS, not from K-1 history. An entity is expected to
  receive a K-1 for tax year Y if Canoe shows it held that fund during Y -- evidenced
  by an Account Statement or Capital Call Notice tagged to that fund and entity, or
  by a K-1 itself. Holding is treated as a contiguous span (first evidence year ->
  last evidence year), so a gap in Canoe's coverage does not punch a hole in the
  expectation. This is what lets the grid flag a fund that has NEVER delivered a
  K-1 -- something K-1 history alone is blind to.

* K-1s are entity-specific by nature, so the grid is entity-grained: two label
  columns (Fund, Entity) rather than the statement tracker's indented sub-rows.

* Deadline is a DATE, not a grace period: June 30 of the year following the tax year
  by default, editable per fund. That is deliberately earlier than the September 15
  extended filing deadline, to leave a chasing window for K-1s still outstanding.

* Two Canoe statuses matter here that do not arise for statements: `Password
  Protected` (the K-1 arrived but is encrypted -- chase the manager for a password)
  gets its own state rather than counting as received, and `Duplicate` documents are
  ignored outright so a re-send never satisfies a year on its own.

Statuses
--------
  received   a clean K-1 covers the year
  late       received, but uploaded after the due date
  pending    expected, not in yet, still before the due date
  OVERDUE    expected, not in, past the due date          <- the chase list
  review     arrived but Canoe flags it for a human
  locked     arrived but password protected
  retag      a K-1 for this fund+year is in Canoe with no entity assigned
  draft      only a DRAFT K-1 has arrived; the final is still owed. Canoe types a draft
             as "K-1" and marks it Complete exactly like the real thing, so the only
             signal is the manager's own filename ("2025 DRAFT K-1 - ...") -- see
             DRAFT_MARKERS. Where a draft and a final both cover a year the final wins
             and the cell is simply received; a draft ALONE must never read as received,
             because a tax return cannot be filed on it.
  unverified holdings say a K-1 is owed and this fund has never delivered one for ANY
             year, so it may simply not issue them (a 1099 payer, an offshore feeder,
             a position held through a blocker). It is painted the same red as OVERDUE
             -- Canoe cannot tell the two apart, so softening it only hid real chases --
             and the difference surfaces as the Chase list's Action column instead:
             "chase the manager" vs "confirm this fund issues a K-1". Retire a genuine
             non-issuer with track=no.

The grid collapses these into FOUR colours (received / not in Canoe / not yet due /
needs a fix). The finer statuses live on in the Chase list and the CSVs, where the
distinction changes what a person does about the row.

Outputs (into <root>/_k1_tracker/ -- team-visible)
  K-1 Tracker <date>.xlsx     THE team workbook, four sheets:
                                "By fund"        the received grid, fund -> entity
                                "By entity"      the same grid pivoted, entity -> fund
                                "Chase list"     flat worklist, most-actionable first
                                "Entity summary" expected/received/outstanding per entity
                              The two grids are the same records grouped two ways --
                              which reads better depends on whether you are asking about
                              a manager or about a tax return. Previous workbooks are
                              swept into Archive/.
  backend/                    K-1 Tracker.html dashboard, k1_status.csv,
                              k1_received_log.csv, K-1 Digest.html, and the editable
                              k1_schedule.xlsx.

Usage:
  python k1_tracker.py --graph                      # unattended job
  python k1_tracker.py --graph --refresh full
  python k1_tracker.py --dest "$CANOE_ARCHIVE_DIR"  # local archive
"""

from __future__ import annotations

import argparse
import collections
import csv
import html
import json
import os
import re
import smtplib
import sys
import time
from datetime import date, datetime, timedelta, timezone
from email.mime.text import MIMEText

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

import config
import statement_tracker as st
from manifest import Manifest
from statement_tracker import (ArchiveLinks, _cell_str, _digest, _free_name,
                               local_inventory, parse_date)

# Documents this tracker is ABOUT. K-3 rides along: it is the foreign-activity
# companion to the K-1 and arrives on the same cycle from the same manager.
K1_TYPE_NAMES = ["K-1", "K-3"]

# Documents that PROVE an entity held a fund in a given year. Both carry an entity
# and a data date; fund-level documents (quarterly reports, financials) are useless
# here because they are addressed to all LPs and carry no entity.
# NB: "Capital Account Statement" and "Distribution Notice" are absent on purpose --
# they return zero documents in this Canoe tenant, so listing them would only widen
# the pull for nothing.
HOLDING_TYPE_NAMES = ["Account Statement", "Capital Call Notice"]

# Canoe statuses that mean a person must look before the year can be trusted.
# Wider than statement_tracker's set: an unconfigured K-1 has no usable allocation,
# so it cannot be allowed to satisfy a year either.
REVIEW_STATUSES = {"awaiting confirmation", "anomaly detected",
                   "potential discrepancy", "configuration required"}

# Arrived, but encrypted -- actionable in a different direction (ask the manager for
# the password) so it gets its own state instead of being lumped into review.
LOCKED_STATUSES = {"password protected"}

# Canoe has already identified these as re-sends of a document it holds. They must
# never satisfy a year by themselves, and they would otherwise double every cell.
IGNORED_STATUSES = {"duplicate"}

# Canoe's couldn't-identify buckets are not funds. Same rule as statement_tracker.
IGNORED_INVESTMENTS = {"unknown", "unknown investment"}
IGNORED_ENTITIES = {"", "--"}

# Default K-1 deadline: June 30 of the year AFTER the tax year. Earlier than the
# Sept 15 extended filing deadline on purpose, to leave a chasing window.
DEFAULT_DUE_MD = "06-30"

# Canoe's own document coverage effectively starts in 2022 (82 K-1s for TY2022
# against 8 for 2018-2021 combined). Seeding expectations before that would paint
# a wall of red for years no-one is going to chase.
DEFAULT_FIRST_TAX_YEAR = 2022

SUBDIR = "_k1_tracker"
BACKEND = "backend"
ARCHIVE = "Archive"
SCHEDULE_FILE = "k1_schedule.xlsx"
CACHE_FILE = "k1_metadata_cache.json"
DIGEST_STATE = "k1_digest_state.json"
# Every run writes a NEW dated workbook and sweeps older ones into Archive/, for the
# same reason statement_tracker does: a fresh file is a fresh OneDrive item, so an
# Excel session holding last week's grid open can never block the update.
GRID_PREFIX = "K-1 Tracker"

# Bump whenever _slim starts keeping a field it did not keep before. A cache written by
# an older build has no such field, and nothing else about the request would change, so
# without this the tracker would happily reconcile against silently incomplete records.
CACHE_SCHEMA = 2

# A manager's own filename is the only place a draft K-1 announces itself: Canoe types
# it "K-1" and marks it Complete exactly like the real thing, document_approval reads
# "Pending" on all 454 documents, and document_tags carry only Canoe-internal flags.
#
# "amended" is deliberately NOT a marker. Every occurrence in this library is the phrase
# "as amended" inside a trust's legal name ("...Trust dated December 23, 1982, as
# amended"), so matching it would flag dozens of perfectly good K-1s.
DRAFT_MARKERS = re.compile(r"\b(draft|prelim\w*|estimated?)\b", re.I)


def is_draft(row_or_doc: dict) -> bool:
    """True when the manager's filename says this is a draft/estimated K-1."""
    return bool(DRAFT_MARKERS.search(row_or_doc.get("original_file_name") or ""))


def _norm(s: str) -> str:
    """Casefold a name for comparison: '&' -> 'and', punctuation and runs of space gone.

    Managers write the entity out in full and in their own style ("Pablo and Tiffany
    Bernal 2016 Trust dated June 3, 2016"), while Canoe stores a short form ("Pablo &
    Tiffany Bernal 2016 Trust"). Normalising both makes the short form a substring.
    """
    s = (s or "").lower().replace("&", " and ")
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]+", " ", s)).strip()


def attribute_untagged(doc: dict, entities, aliases: dict | None = None) -> str | None:
    """Which of `entities` a Canoe-untagged K-1 belongs to, from the manager's filename.

    Canoe leaves an entity blank when its extraction is unsure, but the manager almost
    always names the partner in the filename. Matching that back is deterministic and
    conservative: EXACTLY one known entity must appear, otherwise this returns None and
    the document stays unattributed. When two candidates both match, the longer name
    wins only if it contains the shorter -- "Scott and Amanda Brooks Family Trust" and
    "Scott J Brooks and Amanda T Brooks" are different entities and must never collapse.
    """
    hay = _norm(doc.get("original_file_name") or "")
    if not hay:
        return None
    # A manager may spell the owner the aliased way, so search the alias spellings too
    # and resolve any hit back to the canonical name.
    search = {e: e for e in entities}
    for alias, canon in (aliases or {}).items():
        if canon in entities:
            search[alias] = canon
    hits = [c for name, c in search.items() if _norm(name) and _norm(name) in hay]
    hits = list(dict.fromkeys(hits))          # dedupe, keep order
    if len(hits) == 1:
        return hits[0]
    if len(hits) > 1:
        longest = max(hits, key=lambda e: len(_norm(e)))
        if all(_norm(e) in _norm(longest) for e in hits):
            return longest          # nested aliases of one name
    return None


# --------------------------------------------------------------------------- #
# Metadata pull (cached + incremental)
# --------------------------------------------------------------------------- #

def _slim(doc: dict) -> dict:
    """Keep only the fields this tracker needs, so the cache stays small.

    Deliberately NOT statement_tracker._slim: that one drops validated_data, which is
    how the three K-1s with no data_date recover their tax year. Allocation nesting is
    flattened the same way (Canoe sometimes returns a list inside the list).
    """
    raw = doc.get("allocations") or []
    flat = []
    for a in raw:
        if isinstance(a, list):
            flat.extend(x for x in a if isinstance(x, dict))
        elif isinstance(a, dict):
            flat.append(a)
    allocs = []
    for a in flat:
        # Only the year is kept from Canoe's extraction blobs -- they also contain
        # figures from the document body, which this tool has no business caching.
        years = {}
        for key in ("validated_data", "extracted_data"):
            y = str(((a.get(key) or {}).get("year")) or "").strip()
            if y:
                years[key] = y
        allocs.append({
            "data_date": a.get("data_date"),
            "investment": a.get("investment"),
            "investment_id": a.get("investment_id"),
            "fund_sponsor": a.get("fund_sponsor"),
            "entity": a.get("entity"),
            "dataset_type": a.get("dataset_type"),
            **years,
        })
    return {
        "id": doc.get("id"),
        "name": doc.get("name"),
        # The manager's own filename. Canoe's `name` is a generated template
        # ("<fund>-<entity>-K-1-<date>") and carries no draft/final marker -- this is
        # the ONLY field that distinguishes a draft K-1 from the real one. See is_draft.
        "original_file_name": doc.get("original_file_name"),
        "document_status": doc.get("document_status"),
        "document_type": doc.get("document_type"),
        "uploaded": doc.get("uploaded"),
        "last_modified": doc.get("last_modified"),
        "allocations": allocs,
    }


def _pull(extra: dict, label: str) -> list[dict]:
    """Page through the metadata endpoint. Same defensive pagination as the statement
    tracker (dedupe by id, stop when a page adds nothing new); shares its retry and
    rate-limit handling via st._fetch_page."""
    by_id: dict[str, dict] = {}
    page = 1
    t0 = time.time()
    while page <= 200:
        resp = st._fetch_page(page, extra)
        batch = resp.json()
        if not isinstance(batch, list) or not batch:
            break
        before = len(by_id)
        for d in batch:
            slim = _slim(d)
            if slim.get("id"):
                by_id[slim["id"]] = slim
        total_pages = resp.headers.get("total_pages")
        print(f"  {label}: page {page}{'/' + total_pages if total_pages else ''}"
              f" (+{len(by_id) - before} new, total {len(by_id)}, {time.time() - t0:.0f}s)")
        if len(by_id) == before:
            break
        if total_pages and page >= int(total_pages):
            break
        page += 1
        remaining = resp.headers.get("X-RateLimit-Remaining")
        if remaining is not None and int(remaining) <= 3:
            print("    near rate limit; pausing 60s...")
            time.sleep(60)
    return list(by_id.values())


def load_metadata(cache_path: str, refresh: str) -> list[dict]:
    """K-1/K-3 documents plus the holdings evidence, from cache + delta (or full pull).

    One cache for both type sets: they come from the same endpoint and the same delta
    window, and splitting them would double the request count for no benefit.
    """
    type_names = K1_TYPE_NAMES + HOLDING_TYPE_NAMES
    types_param = ",".join(sorted(type_names))
    cache = None
    if refresh != "full" and os.path.exists(cache_path):
        try:
            cache = json.load(open(cache_path))
        except (OSError, ValueError):
            cache = None
    if cache is not None and set(cache.get("types", [])) != {t.lower() for t in type_names}:
        print("  document-type set changed -- full re-pull")
        cache = None
    if cache is not None and cache.get("schema") != CACHE_SCHEMA:
        print("  cache schema changed (new fields needed) -- full re-pull")
        cache = None
    # A delta by last-modified cannot notice a document re-typed OUT of the tracked
    # set (the type filter simply stops returning it), so re-baseline monthly.
    if cache is not None:
        full_at = cache.get("full_pulled_at")
        if not full_at or (datetime.now(timezone.utc)
                           - datetime.fromisoformat(full_at)).days >= 30:
            print("  monthly re-baseline -- full re-pull")
            cache = None

    if cache is None:
        print("  full metadata pull (first run or --refresh full) -- this can take a while...")
        docs = _pull({"document_type": types_param}, "full pull")
        by_id = {d["id"]: d for d in docs if d.get("id")}
        full_pulled_at = datetime.now(timezone.utc).isoformat()
    else:
        by_id = {d["id"]: d for d in cache.get("docs", []) if d.get("id")}
        # Two days of overlap so a fund/date correction made inside Canoe is picked up.
        since = (datetime.fromisoformat(cache["pulled_at"]) - timedelta(days=2)).strftime("%Y-%m-%d")
        fresh = _pull({"document_type": types_param,
                       "last_modified_time_start": since}, f"delta since {since}")
        for d in fresh:
            if d.get("id"):
                by_id[d["id"]] = d
        print(f"  cache: {len(by_id)} docs after merging {len(fresh)} modified")
        full_pulled_at = cache.get("full_pulled_at") or datetime.now(timezone.utc).isoformat()

    json.dump({"pulled_at": datetime.now(timezone.utc).isoformat(),
               "full_pulled_at": full_pulled_at,
               "schema": CACHE_SCHEMA,
               "types": sorted(t.lower() for t in type_names),
               "docs": list(by_id.values())}, open(cache_path, "w"))
    return list(by_id.values())


# --------------------------------------------------------------------------- #
# Rows
# --------------------------------------------------------------------------- #

def tax_year(a: dict) -> int | None:
    """The tax year an allocation covers.

    data_date is authoritative (12-31 on 445 of 448 K-1s). For the handful Canoe has
    not dated, fall back to the year it extracted from the document -- validated_data
    agrees with data_date on 428 of 448, so it is a trustworthy stand-in. Values come
    through as either "2022" or "Y2022".
    """
    dd = parse_date(a.get("data_date"))
    if dd:
        return dd.year
    for key in ("validated_data", "extracted_data"):
        y = str(a.get(key) or "").strip().lstrip("Yy")
        if re.fullmatch(r"(19|20)\d\d", y):
            return int(y)
    return None


def _split(docs: list[dict], aliases: dict | None = None) -> tuple[list[dict], list[dict], list[dict]]:
    """One pull, three row sets: K-1/K-3 rows, holdings evidence, and unassigned K-1s.

    The third exists because Canoe's "unknown" investment bucket would otherwise swallow
    real K-1s silently. Those documents cannot be placed on the grid -- there is no fund
    to put them under -- but they are K-1s somebody is waiting on, so they are reported
    separately rather than dropped.
    """
    k1_types = {t.lower() for t in K1_TYPE_NAMES}
    hold_types = {t.lower() for t in HOLDING_TYPE_NAMES}
    # Normalise the keys here rather than trusting every caller to pre-lowercase them.
    aliases = {a.lower(): c for a, c in (aliases or {}).items()}
    k1_rows, hold_rows, unassigned = [], [], []
    ignored = 0
    for d in docs:
        dtype = (d.get("document_type") or "").strip()
        status = (d.get("document_status") or "").strip()
        for a in d.get("allocations") or []:
            inv = (a.get("investment") or "").strip()
            if not inv or inv.lower() in IGNORED_INVESTMENTS:
                if dtype.lower() in k1_types and status.lower() not in IGNORED_STATUSES:
                    unassigned.append({
                        "document_type": dtype, "document_status": status,
                        "tax_year": tax_year(a), "doc_id": d.get("id"),
                        "doc_name": d.get("name") or "",
                        "uploaded": parse_date(d.get("uploaded")),
                    })
                continue
            ent = (a.get("entity") or "").strip()
            # Applied here, at the single point rows are built, so holdings, spans,
            # reconciliation, attribution and every sheet all see the same name.
            if aliases:
                ent = aliases.get(ent.lower(), ent)
            year = tax_year(a)
            if dtype.lower() in k1_types or (a.get("dataset_type") or "").strip().lower() in k1_types:
                if status.lower() in IGNORED_STATUSES:
                    ignored += 1
                    continue
                k1_rows.append({
                    "investment": inv,
                    "investment_id": a.get("investment_id") or "",
                    "fund_sponsor": (a.get("fund_sponsor") or "").strip(),
                    "entity": ent,
                    "tax_year": year,
                    "uploaded": parse_date(d.get("uploaded")),
                    "document_type": dtype,
                    "document_status": status,
                    "is_draft": is_draft(d),
                    # Kept on the row, not just consumed here: reconcile needs it to
                    # attribute an entity-less K-1 from the manager's filename.
                    "original_file_name": d.get("original_file_name") or "",
                    "doc_id": d.get("id"),
                    "doc_name": d.get("name") or "",
                })
            elif dtype.lower() in hold_types and ent not in IGNORED_ENTITIES and year:
                # Holdings evidence needs nothing but "this entity held this fund in
                # this year" -- status is irrelevant, a flagged statement still proves
                # the position existed.
                hold_rows.append({"investment": inv, "entity": ent, "year": year,
                                  "fund_sponsor": (a.get("fund_sponsor") or "").strip()})
    if ignored:
        print(f"  ignored     : {ignored} Canoe-flagged duplicate K-1/K-3 allocations")
    return k1_rows, hold_rows, unassigned


def holding_spans(k1_rows: list[dict], hold_rows: list[dict]) -> dict:
    """(investment, entity) -> (first_year, last_year) the entity is shown holding.

    A contiguous span, not the raw set of observed years: Canoe's coverage has gaps,
    and an entity that holds a fund in 2022 and 2025 unquestionably held it in 2023
    and 2024 too. A K-1 is itself proof of holding, so it counts as evidence -- that
    keeps a fund whose only Canoe presence is its K-1s on the grid.
    """
    years: dict[tuple, set] = {}
    for r in hold_rows:
        years.setdefault((r["investment"], r["entity"]), set()).add(r["year"])
    for r in k1_rows:
        if r["entity"] not in IGNORED_ENTITIES and r["tax_year"]:
            years.setdefault((r["investment"], r["entity"]), set()).add(r["tax_year"])
    return {k: (min(v), max(v)) for k, v in years.items() if v}


# --------------------------------------------------------------------------- #
# Schedule (editable config, auto-seeded)
# --------------------------------------------------------------------------- #

SCHEDULE_HEADER = ["investment", "fund_sponsor", "contact", "track", "first_tax_year",
                   "due_month_day", "exclude_entities", "notes"]

# Entity aliases live on their own sheet, not in a schedule column: one real owner can be
# spelled two ways across MANY funds, so the mapping is global rather than per-fund.
ALIAS_SHEET = "Entity aliases"
ALIAS_HEADER = ["alias", "canonical_entity", "notes"]
ALIAS_HELP = [
    "Entity aliases -- collapse two Canoe entity names that are really the same owner.",
    "",
    "alias           : the Canoe entity name to replace (exactly as Canoe spells it).",
    "canonical_entity: the name to use instead. Every K-1, statement and capital call",
    "                  tagged to `alias` is counted against `canonical_entity`, in every",
    "                  fund, so the two stop appearing as separate rows on the grid.",
    "",
    "Use this when Canoe tags one owner inconsistently -- e.g. the same position filed",
    "sometimes under a trust's name and sometimes under the trustees' personal names.",
    "",
    "Do NOT use it to merge genuinely different owners. Two entities that file separate",
    "tax returns must stay separate rows, or the grid will report one as complete when",
    "the other's K-1 is still outstanding.",
    "",
    "Matching ignores case. An alias pointing at itself is ignored.",
]

SCHEDULE_HELP = [
    "K-1 tracker schedule -- edit freely; the tracker re-reads this file each run.",
    "",
    "contact         : who to chase for this fund's K-1 (name / email). Free text,",
    "                  carried straight through to the Chase list sheet. Worth filling",
    "                  in: Canoe leaves fund_sponsor empty on K-1 allocations, so this",
    "                  is the only place the workbook can learn who to email.",
    "track           : yes | no  (no = fund ignored entirely, e.g. a wound-down fund",
    "                  whose final K-1 is already in).",
    "first_tax_year  : earliest tax year to expect a K-1 for (blank = "
    f"{DEFAULT_FIRST_TAX_YEAR}).",
    f"due_month_day   : deadline as MM-DD in the year AFTER the tax year (blank = "
    f"{DEFAULT_DUE_MD}).",
    "                  Default is deliberately ahead of the Sept 15 extended filing",
    "                  deadline, to leave time to chase. Push it out per fund for a",
    "                  manager who reliably files on extension.",
    "exclude_entities: ;-separated entity names that never receive a K-1 for this",
    "                  fund (e.g. a position held through a blocker, or an entity",
    "                  whose K-1 arrives inside a composite return).",
    "",
    "WHICH YEARS ARE EXPECTED is derived from holdings, not from this file: an entity",
    "is expected to get a K-1 for a tax year when Canoe shows it held the fund that",
    "year (Account Statement / Capital Call Notice / a K-1). Use first_tax_year and",
    "exclude_entities to trim; there is nothing to add by hand.",
    "",
    "New funds appearing in Canoe are appended automatically with a NEW note.",
]


def seed_schedule(k1_rows: list[dict], hold_rows: list[dict], spans: dict,
                  path: str, aliases: dict | None = None) -> list[dict]:
    """First run: one row per fund, derived from what Canoe already shows."""
    sponsors: dict[str, list[str]] = {}
    for r in k1_rows + hold_rows:
        if r.get("fund_sponsor"):
            sponsors.setdefault(r["investment"], []).append(r["fund_sponsor"])

    funds = {r["investment"] for r in k1_rows} | {r["investment"] for r in hold_rows}
    sched = []
    for inv in sorted(funds, key=str.lower):
        sp = sponsors.get(inv, [])
        sponsor = max(set(sp), key=sp.count) if sp else ""
        ent_spans = {e: s for (i, e), s in spans.items() if i == inv}
        n_k1 = sum(1 for r in k1_rows if r["investment"] == inv)
        first = min((s[0] for s in ent_spans.values()), default=None)
        last = max((s[1] for s in ent_spans.values()), default=None)
        notes = (f"auto-seeded: {len(ent_spans)} entity(ies), {n_k1} K-1/K-3 on file"
                 + (f", holdings {first}-{last}" if first else ""))
        if n_k1 == 0:
            notes += " -- NO K-1 EVER RECEIVED, verify this fund issues one"
        sched.append({
            "investment": inv,
            "fund_sponsor": sponsor,
            "contact": "",
            "track": "yes" if ent_spans else "no",
            "first_tax_year": str(max(first, DEFAULT_FIRST_TAX_YEAR)) if first else "",
            "due_month_day": "",
            "exclude_entities": "",
            "notes": notes,
        })
    write_schedule(path, sched, aliases)
    return sched


def load_entity_aliases(path: str) -> dict:
    """{alias_lowercased: canonical} from the schedule workbook's alias sheet."""
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[str, str] = {}
    if ALIAS_SHEET in wb.sheetnames:
        rows = wb[ALIAS_SHEET].iter_rows(values_only=True)
        header = [_cell_str(h) for h in next(rows, [])]
        for r in rows:
            rec = {h: _cell_str(v) for h, v in zip(header, r) if h}
            alias, canon = rec.get("alias", ""), rec.get("canonical_entity", "")
            if alias and canon and alias.strip().lower() != canon.strip().lower():
                out[alias.strip().lower()] = canon.strip()
    wb.close()
    return out


def write_schedule(path: str, sched: list[dict], aliases: dict | None = None) -> None:
    """Write the schedule workbook, preserving the alias sheet.

    `aliases` must always be passed the ones currently in force: this rewrites the whole
    file, so omitting them would silently discard hand-entered rows the next time a new
    fund is appended.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.append(SCHEDULE_HEADER)
    for c in ws[1]:
        c.font = Font(bold=True)
    for s in sched:
        ws.append([s.get(k, "") for k in SCHEDULE_HEADER])
    last = max(len(sched) + 200, 500)
    dv_track = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    ws.add_data_validation(dv_track)
    col_t = SCHEDULE_HEADER.index("track") + 1
    letter = openpyxl.utils.get_column_letter(col_t)
    dv_track.add(f"{letter}2:{letter}{last}")
    for col, w in {"A": 42, "B": 22, "C": 30, "D": 7, "E": 14,
                   "F": 14, "G": 40, "H": 66}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    al = wb.create_sheet(ALIAS_SHEET)
    al.append(ALIAS_HEADER)
    for c in al[1]:
        c.font = Font(bold=True)
    for alias, canon in sorted((aliases or {}).items()):
        al.append([alias, canon, ""])
    for col, w in {"A": 46, "B": 46, "C": 50}.items():
        al.column_dimensions[col].width = w
    al.freeze_panes = "A2"
    start = len(aliases or {}) + 3
    for i, line in enumerate(ALIAS_HELP):
        al.cell(start + i, 1, line).font = Font(italic=True, color="666666")

    info = wb.create_sheet("How to use")
    for line in SCHEDULE_HELP:
        info.append([line])
    info.column_dimensions["A"].width = 100
    wb.save(path)


def load_schedule(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Schedule"] if "Schedule" in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    header = [_cell_str(h) for h in next(rows, [])]
    out = []
    for r in rows:
        rec = {h: _cell_str(v) for h, v in zip(header, r) if h}
        if rec.get("investment"):
            out.append(rec)
    wb.close()
    return out


def sync_new_funds(sched: list[dict], k1_rows: list[dict], hold_rows: list[dict],
                   spans: dict, path: str, aliases: dict | None = None) -> list[dict]:
    """Funds in Canoe but not yet in the schedule get appended, flagged NEW."""
    known = {s["investment"] for s in sched}
    seen = {r["investment"] for r in k1_rows} | {r["investment"] for r in hold_rows}
    new_invs = sorted(seen - known, key=str.lower)
    if not new_invs:
        return sched
    seeded = seed_schedule(k1_rows, hold_rows, spans, path + ".tmp", aliases)
    os.remove(path + ".tmp")
    by_inv = {s["investment"]: s for s in seeded}
    for inv in new_invs:
        s = by_inv.get(inv)
        if s:
            s["notes"] = "NEW -- " + s["notes"]
            sched.append(s)
            print(f"  new fund added to schedule: {inv}")
    write_schedule(path, sched, aliases)
    return sched


# --------------------------------------------------------------------------- #
# Reconciliation
# --------------------------------------------------------------------------- #

# A real K-1 beats its foreign-activity companion when both cover a year.
TYPE_PRIORITY = {"k-1": 0, "k-3": 1}


def _doc_rank(m: dict) -> tuple:
    """Sort key picking which document represents a cell: final before draft, K-1 before
    K-3, then oldest first. Managers routinely send a draft and then the real thing, so
    the final must win the cell's link even though the draft arrived first."""
    return (bool(m.get("is_draft")),
            TYPE_PRIORITY.get((m["document_type"] or "").lower(), 9),
            m["uploaded"] or date.min)


def due_date(year: int, due_md: str) -> date:
    """The deadline for a tax year: MM-DD in the FOLLOWING calendar year.

    Anything unparseable falls back to the default rather than raising: due_month_day
    is hand-edited in a workbook, and a typo there must not take down the unattended
    job. AttributeError covers a blank cell arriving as None.
    """
    for candidate in (due_md, DEFAULT_DUE_MD):
        try:
            month, day = (int(x) for x in candidate.split("-"))
            return date(year + 1, month, day)
        except (AttributeError, ValueError, TypeError):
            continue
    raise ValueError(f"DEFAULT_DUE_MD is not a valid MM-DD: {DEFAULT_DUE_MD!r}")


def latest_tax_year(today: date) -> int:
    """The most recent tax year a K-1 could exist for: last calendar year."""
    return today.year - 1


def reconcile(sched: list[dict], k1_rows: list[dict], spans: dict,
              today: date, aliases: dict | None = None) -> list[dict]:
    """One record per expected (fund, entity, tax year), with a status.

    Rows exist only where holdings say a K-1 is owed, so a green cell means "arrived"
    and a red one means "owed and missing" -- there is no third reading.
    """
    max_year = latest_tax_year(today)

    # Funds that have delivered at least one K-1/K-3 for SOME year, entity-tagged or
    # not. A fund absent from this set has never produced one at all, so a missing
    # year there is a classification question, not a manager to chase.
    funds_with_k1 = {r["investment"] for r in k1_rows}

    # K-1s Canoe has not tagged to an entity, by (fund, year): a missing cell with one
    # of these present is a re-tagging job in Canoe, not a missing K-1 from a manager.
    #
    # Each is first attributed to an entity from the manager's filename where that is
    # unambiguous. Without this the same untagged document is offered to every entity
    # missing that fund-year, which sends one beneficiary to another's tax return.
    entities_by_fund: dict[str, set] = {}
    for (i, e) in spans:
        entities_by_fund.setdefault(i, set()).add(e)

    untagged: dict[tuple, list[dict]] = {}
    attributed: dict[tuple, list[dict]] = {}
    n_attributed = 0
    for r in k1_rows:
        if r["entity"] in IGNORED_ENTITIES and r["tax_year"]:
            owner = attribute_untagged(r, entities_by_fund.get(r["investment"], ()), aliases)
            if owner:
                attributed.setdefault((r["investment"], owner, r["tax_year"]), []).append(r)
                n_attributed += 1
            else:
                untagged.setdefault((r["investment"], r["tax_year"]), []).append(r)
    for v in list(untagged.values()) + list(attributed.values()):
        v.sort(key=_doc_rank)
    if n_attributed:
        print(f"  attributed  : {n_attributed} entity-less K-1/K-3 matched to an entity "
              f"by the manager's filename")

    by_cell: dict[tuple, list[dict]] = {}
    for r in k1_rows:
        if r["entity"] not in IGNORED_ENTITIES and r["tax_year"]:
            by_cell.setdefault((r["investment"], r["entity"], r["tax_year"]), []).append(r)

    out = []
    for s in sched:
        if (s.get("track") or "").strip().lower() not in ("yes", "y", "true", "1"):
            continue
        inv = s["investment"]
        if inv.strip().lower() in IGNORED_INVESTMENTS:
            continue
        try:
            floor = int(s.get("first_tax_year") or DEFAULT_FIRST_TAX_YEAR)
        except ValueError:
            floor = DEFAULT_FIRST_TAX_YEAR
        due_md = (s.get("due_month_day") or "").strip() or DEFAULT_DUE_MD
        excluded = {e.strip().lower() for e in (s.get("exclude_entities") or "").split(";")
                    if e.strip()}

        for (i, entity), (first, last) in sorted(spans.items(), key=lambda kv: kv[0][1].lower()):
            if i != inv or entity.lower() in excluded:
                continue
            for year in range(max(first, floor), min(last, max_year) + 1):
                due = due_date(year, due_md)
                matched = sorted(by_cell.get((inv, entity, year), []), key=_doc_rank)
                clean = [m for m in matched
                         if m["document_status"].lower() not in REVIEW_STATUSES
                         and m["document_status"].lower() not in LOCKED_STATUSES]
                flagged = [m for m in matched
                           if m["document_status"].lower() in REVIEW_STATUSES]
                locked = [m for m in matched
                          if m["document_status"].lower() in LOCKED_STATUSES]
                if clean:
                    best = clean[0]
                    final = [m for m in clean if not m.get("is_draft")]
                    if not final:
                        # Only a draft has arrived. It cannot be filed on, so the year is
                        # still open -- counting it as received would tell the tax
                        # preparer they have a K-1 they cannot actually use.
                        status, received = "draft", None
                    else:
                        received = min((m["uploaded"] for m in final if m["uploaded"]),
                                       default=None)
                        status = "received" if (received is None or received <= due) else "late"
                elif flagged:
                    # Internal Canoe work comes before chasing a password.
                    status, received, best = "review", None, flagged[0]
                elif locked:
                    status, received, best = "locked", None, locked[0]
                elif attributed.get((inv, entity, year)):
                    # The filename names THIS entity, so the link is safe to offer.
                    status, received, best = "retag", None, attributed[(inv, entity, year)][0]
                elif untagged.get((inv, year)):
                    status, received, best = "retag", None, untagged[(inv, year)][0]
                elif inv not in funds_with_k1:
                    status, received, best = "unverified", None, None
                else:
                    status, received, best = ("overdue" if today > due else "pending"), None, None
                out.append({
                    # Canoe leaves fund_sponsor empty on K-1 allocations, so `contact`
                    # from the schedule is what actually tells the chase list who to email.
                    "fund_sponsor": s.get("fund_sponsor") or "",
                    "contact": s.get("contact") or "",
                    "investment": inv,
                    "entity": entity,
                    "tax_year": year,
                    "due": due.isoformat(),
                    "days_late": (today - due).days if (today > due and status in CHASE_STATUSES) else "",
                    "status": status,
                    "received_date": received.isoformat() if received else "",
                    "document_type": best["document_type"] if best else "",
                    "document_status": best["document_status"] if best else "",
                    "doc_name": best["doc_name"] if best else "",
                    "doc_id": best["doc_id"] if best else "",
                    "n_docs": len(matched),
                    # Final (non-draft, un-flagged) K-1s on this cell. Two of them for
                    # one entity while a sibling entity has none is the signature of a
                    # mis-tag -- see the pass below.
                    "n_final": sum(1 for m in clean if not m.get("is_draft")),
                    "n_untagged": len(untagged.get((inv, year), [])),
                    # Every record carries both keys with the same shape. Setting them
                    # only on the ambiguous ones made the record type heterogeneous, and
                    # a DictWriter that takes its columns from the first row then blew up
                    # on a later one.
                    "ambiguous": False,
                    "n_claimants": 1,
                    "mistag_check": False,
                    "mistag_hint": "",
                })

    # A single untagged K-1 cannot belong to two entities at once. The retag rule above
    # offers the same document to every entity row missing that fund-year, which is a
    # useful nudge when exactly one row is missing and a dangerous lie when several are:
    # it points one beneficiary at another beneficiary's tax document. Mark those cells
    # ambiguous so the grid and the Chase list stop implying ownership.
    competing: dict[tuple, int] = {}
    for r in out:
        if r["status"] == "retag":
            competing[(r["investment"], r["tax_year"])] = \
                competing.get((r["investment"], r["tax_year"]), 0) + 1
    for r in out:
        if r["status"] == "retag" and competing[(r["investment"], r["tax_year"])] > 1:
            r["ambiguous"] = True
            r["n_claimants"] = competing[(r["investment"], r["tax_year"])]

    # Mis-tag detection. Within one fund-year, an entity holding TWO final K-1s while a
    # sibling entity holds none is the signature of a mis-tag: a manager sends one K-1
    # per partner, so the surplus one probably belongs to the entity showing nothing.
    # Flagging both sides matters because the missing row otherwise reads "chase the
    # manager" for a document already sitting in Canoe under the wrong name.
    #
    # A suspicion, not a verdict -- an amended K-1 also lands as a second final -- so it
    # asks for a check rather than reassigning anything.
    groups: dict[tuple, list[dict]] = {}
    for r in out:
        groups.setdefault((r["investment"], r["tax_year"]), []).append(r)
    for group in groups.values():
        surplus = [r for r in group if r["n_final"] >= 2]
        missing = [r for r in group if r["status"] in ("overdue", "unverified")]
        if not (surplus and missing):
            continue
        hint = "; ".join(f'{r["entity"]} has {r["n_final"]}' for r in surplus)
        for r in missing:
            r["mistag_check"] = True
            r["mistag_hint"] = hint
        for r in surplus:
            r["mistag_check"] = True
            r["mistag_hint"] = "; ".join(f'{r["entity"]} has none' for r in missing)
    return out


# --------------------------------------------------------------------------- #
# Outputs
# --------------------------------------------------------------------------- #

STATUS_META = {
    # status -> (label, pill background, pill text)
    "received": ("Received", "#dcefdc", "#1b5e20"),
    "late":     ("Received late", "#e3ecdf", "#33691e"),
    "pending":  ("Pending", "#eaf0fa", "#1f4e79"),
    "overdue":  ("OVERDUE", "#fbe0dd", "#b71c1c"),
    "review":   ("Review", "#e8e0f5", "#4a148c"),
    "locked":   ("Password protected", "#fce6d4", "#8a4b16"),
    "retag":    ("Needs entity tag", "#fdf3d7", "#8a6d1a"),
    "unverified": ("Not in Canoe", "#fbe0dd", "#b71c1c"),
    "draft":    ("Draft only", "#fdf3d7", "#8a6d1a"),
}

# Chasing a manager for a known-missing K-1. A draft-only year belongs here: the final
# is still outstanding and only the manager can supply it.
ACTION_STATUSES = ("overdue", "draft", "review", "locked", "retag")
# Everything that wants a human. Order is most-actionable-first and drives the sort in
# both the workbook and the dashboard.
CHASE_STATUSES = ACTION_STATUSES + ("unverified",)

# The grid collapses the eight statuses into FOUR colours. The finer statuses still
# drive the Chase list -- where the difference between chasing a manager and confirming
# a fund even issues a K-1 is a different phone call -- but a colour legend is the wrong
# place for that nuance, and an eight-colour key reads as noise.
#
# In particular `unverified` is painted the same red as `overdue`: Canoe cannot tell
# whether a fund issues K-1s (investment_structure says drawdown_fund for 11 of the 13
# funds that have never filed one, exactly as it does for the 50 that have), so treating
# "never seen a K-1 here" as a softer state hid genuine chases -- Bow River II being the
# case in point. Red means "expected and not in Canoe"; the Chase list says what to do.
GRID_STATE = {
    "received": "received", "late": "received",
    "pending": "pending",
    "overdue": "outstanding", "unverified": "outstanding",
    "review": "attention", "locked": "attention", "retag": "attention",
    "draft": "attention",
}

# What the cell says when it is worth clicking. Blank states carry colour only.
CELL_TEXT = {"received": "Link", "late": "Link",
             "review": "Review", "locked": "Locked", "retag": "Tag", "draft": "Draft"}


def cell_label(rec: dict) -> str | None:
    """The clickable text for a cell, or None for a colour-only cell.

    "Tag?" rather than "Tag" when several entities are competing for the same untagged
    K-1: the document behind the link belongs to at most one of them, so the question
    mark is the difference between a hint and a false claim.
    """
    if rec["status"] == "retag" and rec.get("ambiguous"):
        return "Tag?"
    # A missing K-1 that another entity may already be holding under the wrong tag:
    # worth a look in Canoe before anyone emails the manager.
    if rec.get("mistag_check") and rec["status"] in ("overdue", "unverified"):
        return "Check?"
    return CELL_TEXT.get(rec["status"])

# The Chase list's verdict column: what a person actually does about this row. This is
# where "chase" and "confirm" part company.
CHASE_ACTION = {
    "overdue":    "Chase the manager",
    "draft":      "Chase the manager for the FINAL",
    "unverified": "Confirm this fund issues a K-1",
    "review":     "Clear the flag in Canoe",
    "locked":     "Request the password",
    "retag":      "Assign the entity in Canoe",
}


def write_status_csv(path: str, recs: list[dict]) -> None:
    """Every reconciled record, one row per line.

    Columns come from the UNION of keys across all records, not from the first one: the
    first record is not guaranteed to be representative, and a status that only some rows
    carry would otherwise abort the whole run at the very last step.
    """
    if not recs:
        return
    fields = list(recs[0].keys())
    for r in recs:
        fields.extend(k for k in r if k not in fields)
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(recs)


def write_received_log(path: str, k1_rows: list[dict]) -> None:
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["investment", "fund_sponsor", "entity", "document_type",
                    "tax_year", "uploaded", "document_status", "doc_name"])
        for r in sorted(k1_rows, key=lambda r: (r["investment"].lower(),
                                                r["tax_year"] or 0,
                                                r["entity"].lower())):
            w.writerow([r["investment"], r["fund_sponsor"], r["entity"],
                        r["document_type"], r["tax_year"] or "",
                        r["uploaded"].isoformat() if r["uploaded"] else "",
                        r["document_status"], r["doc_name"]])


def _write_grid(ws, recs: list[dict], links: ArchiveLinks, primary: str, secondary: str,
                headers: tuple[str, str], widths: tuple[int, int], style: dict) -> None:
    """One received-grid, grouped by `primary` then `secondary`, a column per tax year.

    Orientation is the only difference between the "By fund" and "By entity" sheets, so
    it is a parameter rather than a second copy of the painting logic -- the two views
    must always agree, and the surest way to guarantee that is one code path.
    """
    fills, text_color = style["fills"], style["text_color"]
    thin, center = style["thin"], style["center"]
    hdr = style["hdr_row"]

    for i, (fill, label) in enumerate(style["legend"], start=1):
        sw = ws.cell(i, 2)
        sw.border = thin
        if fill:
            sw.fill = fill
        ws.cell(i, 3, label)
    ws.cell(len(style["legend"]) + 1, 3, style["note"]).font = Font(italic=True, color="666666")

    years = sorted({r["tax_year"] for r in recs})
    ws.cell(hdr, 1, headers[0]).font = Font(bold=True)
    ws.cell(hdr, 2, headers[1]).font = Font(bold=True)
    for j, y in enumerate(years, start=3):
        c = ws.cell(hdr, j, f"TY {y}")
        c.font = Font(bold=True)
        c.alignment = center

    cell_rec = {(r[primary], r[secondary], r["tax_year"]): r for r in recs}
    i = hdr + 1
    for key in sorted({r[primary] for r in recs}, key=str.lower):
        subs = sorted({r[secondary] for r in recs if r[primary] == key}, key=str.lower)
        for n, sub in enumerate(subs):
            # The group name prints once, on its first sub-row, so the eye groups them
            # without needing merged cells (which break sorting and filtering).
            gc = ws.cell(i, 1, key if n == 0 else "")
            gc.border = thin
            if n == 0:
                gc.font = Font(bold=True)
            sc = ws.cell(i, 2, sub)
            sc.border = thin
            sc.font = Font(color="444444")
            for j, y in enumerate(years, start=3):
                c = ws.cell(i, j)
                c.border = thin
                rec = cell_rec.get((key, sub, y))
                if rec is None:
                    continue
                state = GRID_STATE[rec["status"]]
                c.fill = fills[state]
                label = cell_label(rec)
                if not label:
                    continue
                c.value = label
                c.alignment = center
                # An ambiguous retag cell gets NO hyperlink: the only candidate document
                # belongs to at most one of the competing entities, so linking it would
                # hand somebody another beneficiary's tax return. The document is still
                # reachable -- it is listed once, per fund-year, in the Chase list.
                link = None if (rec["status"] == "retag" and rec.get("ambiguous")) \
                    else links.url(rec)
                if link:
                    c.hyperlink = link
                    c.font = Font(color=text_color.get(state, "000000"), underline="single")
                else:
                    c.font = Font(color=text_color.get(state, "000000"))
            i += 1

    ws.column_dimensions["A"].width = widths[0]
    ws.column_dimensions["B"].width = widths[1]
    for j in range(3, 3 + len(years)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 11
    ws.freeze_panes = f"C{hdr + 1}"


def write_xlsx(path: str, recs: list[dict], links: ArchiveLinks, today: date) -> None:
    """The team workbook: the received grid in both orientations (By fund / By entity),
    the Chase list worklist, and the per-entity Entity summary.

    Unlike the statement grid, pending and outstanding get different colours. With a June
    deadline the current tax year is legitimately outstanding for half the year, and
    painting that red would train people to ignore red.
    """
    fills = {
        "received":    PatternFill("solid", fgColor="63BE7B"),
        "pending":     PatternFill("solid", fgColor="DDEBF7"),
        "outstanding": PatternFill("solid", fgColor="F8696B"),
        "attention":   PatternFill("solid", fgColor="FFD966"),
    }
    # One entry per grid state: an outstanding cell now carries text too ("Check?"), and
    # a missing colour here is a hard crash rather than a cosmetic slip.
    text_color = {"received": "1B5E20", "attention": "7F6000",
                  "outstanding": "9C0006", "pending": "1F4E79"}
    thin = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    center = Alignment(horizontal="center")

    # Four colours, five lines. NB: a label must not start with "=" or Excel reads it as
    # a formula (#NAME?).
    LEGEND = [
        (fills["received"], 'Received -- click "Link" to open the K-1'),
        (fills["outstanding"], 'Not in Canoe -- see the Chase list sheet. "Check?" means '
                               'another entity holds 2+ K-1s for this fund and year, so one '
                               'of them may be mis-tagged and actually be this one.'),
        (fills["pending"], "Not yet due"),
        (fills["attention"], 'In Canoe but not usable yet -- "Draft" (final still owed), '
                             '"Review", "Locked", "Tag". Click to open it. "Tag?" means an '
                             'untagged K-1 exists for the fund/year but several entities '
                             'are missing it, so it belongs to only one of them.'),
        (None, "Not expected -- the entity did not hold this fund that year"),
    ]
    NOTE = ("A row exists only where Canoe shows the entity held the fund that year, so an "
            "empty cell means no K-1 is owed. The Chase list sheet says what to do about "
            f"every red and amber cell; edit {SCHEDULE_FILE} to change a fund's deadline, "
            "earliest year, excluded entities, or to stop tracking it.")
    HDR_ROW = len(LEGEND) + 3
    style = {"fills": fills, "text_color": text_color, "thin": thin, "center": center,
             "legend": LEGEND, "note": NOTE, "hdr_row": HDR_ROW}

    wb = openpyxl.Workbook()
    # The same records, grouped two ways. Which one reads better depends on the question:
    # "has this manager sent everything?" wants funds, "can this entity file yet?" wants
    # entities. Both are one call to _write_grid, so they can never drift apart.
    ws = wb.active
    ws.title = "By fund"
    _write_grid(ws, recs, links, "investment", "entity", ("Fund", "Entity"), (40, 42), style)
    _write_grid(wb.create_sheet("By entity"), recs, links, "entity", "investment",
                ("Entity", "Fund"), (42, 40), style)

    # -- Chase list: the flat worklist behind the June deadline ---------------
    chase = [r for r in recs if r["status"] in CHASE_STATUSES]
    chase.sort(key=lambda r: (CHASE_STATUSES.index(r["status"]), -r["tax_year"],
                              r["investment"].lower(), r["entity"].lower()))
    cs = wb.create_sheet("Chase list")
    cs.cell(1, 1, f"K-1s outstanding as of {today.isoformat()}").font = Font(bold=True, size=13)
    cs.cell(2, 1, "Sorted most-actionable first. The Action column is the point: a fund that "
                  "has filed K-1s before is a manager to chase, whereas one that never has "
                  "may not issue them at all -- confirm before chasing, then set track=no in "
                  "the schedule to retire the row for good.").font = Font(italic=True, color="666666")
    head = ["Action", "Status", "Fund", "Chase who", "Entity", "Tax year", "Due",
            "Days past due", "Detail"]
    for j, h in enumerate(head, start=1):
        c = cs.cell(4, j, h)
        c.font = Font(bold=True)
        c.border = thin
    for n, r in enumerate(chase, start=5):
        detail = {
            "overdue": "no K-1 in Canoe",
            "review": f'{r["document_type"]} present, Canoe status: {r["document_status"]}',
            "locked": "K-1 present but encrypted",
            "retag": (
                f'{r["n_untagged"]} untagged K-1 for this fund/year but '
                f'{r.get("n_claimants")} entities are missing it -- the linked document '
                f'belongs to only ONE of them; open it to see which, then assign in Canoe'
                if r.get("ambiguous") else
                f'K-1 present, no entity assigned: {r["doc_name"]}'),
            "unverified": "no K-1 from this fund for any year",
            "draft": "only a DRAFT K-1 has arrived -- the final is still outstanding",
        }[r["status"]]
        if r.get("mistag_check") and r["status"] in ("overdue", "unverified"):
            # Do not send anyone to the manager for a document Canoe already holds.
            action = "CHECK CANOE TAGGING first"
            detail = (f'{r["mistag_hint"]} K-1 for this fund/year -- one may be mis-tagged '
                      f'and actually belong here. Verify in Canoe before chasing.')
        else:
            action = CHASE_ACTION[r["status"]]
        vals = [action, STATUS_META[r["status"]][0], r["investment"],
                r["contact"] or r["fund_sponsor"], r["entity"], r["tax_year"],
                r["due"], r["days_late"], detail]
        for j, v in enumerate(vals, start=1):
            c = cs.cell(n, j, v)
            c.border = thin
            if j == 2:
                c.fill = fills[GRID_STATE[r["status"]]]
        link = None
        if r["status"] in ("review", "locked", "draft") or \
                (r["status"] == "retag" and not r.get("ambiguous")):
            link = links.url(r)
        if link:
            c = cs.cell(n, 9)
            c.hyperlink = link
            c.font = Font(color="0563C1", underline="single")
    # The unattributable untagged documents, listed ONCE each rather than repeated under
    # every entity that might own them. This is the only place they are linked, so the
    # document is still one click away without any cell claiming it.
    seen_docs: dict[tuple, dict] = {}
    for r in recs:
        if r["status"] == "retag" and r.get("ambiguous") and r["doc_id"] not in seen_docs:
            seen_docs[r["doc_id"]] = r
    n = cs.max_row + 2
    if seen_docs:
        cs.cell(n, 1, "Untagged K-1s -- assign an entity in Canoe, then these resolve "
                      "themselves").font = Font(bold=True, size=12)
        n += 1
        for j, h in enumerate(["Fund", "Tax year", "Candidates", "Document"], start=1):
            c = cs.cell(n, j, h)
            c.font = Font(bold=True)
            c.border = thin
        n += 1
        for r in sorted(seen_docs.values(), key=lambda r: (r["investment"], -r["tax_year"])):
            vals = [r["investment"], r["tax_year"],
                    f'{r.get("n_claimants")} entities missing this year', r["doc_name"]]
            for j, v in enumerate(vals, start=1):
                c = cs.cell(n, j, v)
                c.border = thin
                if j == 1:
                    c.fill = fills["attention"]
            link = links.url(r)
            if link:
                c = cs.cell(n, 4)
                c.hyperlink = link
                c.font = Font(color="0563C1", underline="single")
            n += 1

    for col, w in {"A": 30, "B": 18, "C": 34, "D": 22, "E": 38, "F": 10, "G": 12,
                   "H": 14, "I": 50}.items():
        cs.column_dimensions[col].width = w
    cs.freeze_panes = "A5"

    write_entity_sheet(wb, recs, thin, center, fills)
    wb.save(path)


def write_entity_sheet(wb, recs: list[dict], thin, center, fills) -> None:
    """One row per investing entity: where does each stand, per tax year and overall.

    K-1s differ from statements in who the answer is for. A statement matters per fund,
    but a K-1 matters per ENTITY -- each one files its own return and cannot file until
    every K-1 it is owed has arrived. This sheet answers that directly: "can this entity
    file for TY2025 yet, and if not, how many are we waiting on."
    """
    ws = wb.create_sheet("Entity summary")
    years = sorted({r["tax_year"] for r in recs})
    entities = sorted({r["entity"] for r in recs}, key=str.lower)

    ws.cell(1, 1, "Where each entity stands").font = Font(bold=True, size=13)
    ws.cell(2, 1, "Per tax year: received / expected. An entity can file once its row reads "
                  "n/n for that year.").font = Font(italic=True, color="666666")

    head = ["Entity"] + [f"TY {y}" for y in years] + \
           ["Expected", "Received", "Not in Canoe", "Needs a fix", "Not yet due", "% complete"]
    for j, h in enumerate(head, start=1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True)
        c.border = thin
        if j > 1:
            c.alignment = center

    for n, ent in enumerate(entities, start=5):
        rows = [r for r in recs if r["entity"] == ent]
        lbl = ws.cell(n, 1, ent)
        lbl.border = thin
        for j, y in enumerate(years, start=2):
            yr = [r for r in rows if r["tax_year"] == y]
            c = ws.cell(n, j)
            c.border = thin
            c.alignment = center
            if not yr:
                continue
            got = sum(1 for r in yr if r["status"] in ("received", "late"))
            c.value = f"{got}/{len(yr)}"
            # Colour the year cell by how complete it is, so a scan down a column shows
            # which entities are ready to file.
            c.fill = fills["received"] if got == len(yr) else (
                fills["pending"] if any(r["status"] == "pending" for r in yr)
                else fills["outstanding"])
        tally = collections.Counter(r["status"] for r in rows)
        got = tally["received"] + tally["late"]
        missing = tally["overdue"] + tally["unverified"]
        needs = tally["review"] + tally["locked"] + tally["retag"]
        for j, v in enumerate([len(rows), got, missing, needs, tally["pending"],
                               f"{got / len(rows):.0%}" if rows else ""],
                              start=2 + len(years)):
            c = ws.cell(n, j, v)
            c.border = thin
            c.alignment = center

    ws.column_dimensions["A"].width = 46
    for j in range(2, 2 + len(years)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 10
    for j in range(2 + len(years), 2 + len(years) + 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 13
    ws.freeze_panes = "B5"


def write_html(path: str, recs: list[dict], undated: list[dict], generated: str,
               unassigned: list[dict] | None = None) -> None:
    counts: dict[str, int] = {k: 0 for k in STATUS_META}
    for r in recs:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    years = sorted({r["tax_year"] for r in recs})
    funds = sorted({(r["fund_sponsor"], r["investment"]) for r in recs},
                   key=lambda t: (t[0].lower() or "~", t[1].lower()))

    def esc(s):
        return html.escape(str(s))

    parts = [f"""<!-- generated by k1_tracker.py -->
<meta charset="utf-8">
<title>K-1 Tracker</title>
<style>
 body {{ font: 14px/1.45 -apple-system, "Segoe UI", Helvetica, Arial, sans-serif;
        margin: 24px; color: #24292f; }}
 h1 {{ font-size: 20px; margin: 0 0 2px; }}
 .sub {{ color: #57606a; margin-bottom: 18px; }}
 .cards {{ display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 20px; }}
 .card {{ border: 1px solid #d0d7de; border-radius: 8px; padding: 10px 16px; min-width: 130px; }}
 .card b {{ display: block; font-size: 22px; }}
 table {{ border-collapse: collapse; margin-bottom: 28px; width: 100%; }}
 th, td {{ border: 1px solid #d8dee4; padding: 4px 8px; text-align: left; white-space: nowrap; }}
 th {{ background: #f6f8fa; position: sticky; top: 0; }}
 td.fund {{ max-width: 300px; overflow: hidden; text-overflow: ellipsis; }}
 .pill {{ display: inline-block; border-radius: 10px; padding: 1px 8px; font-size: 12px; }}
 .sponsor {{ color: #57606a; font-size: 12px; }}
 h2 {{ font-size: 16px; margin: 24px 0 8px; }}
 .overdue-row td {{ background: #fff5f4; }}
</style>
<h1>K-1 Tracker</h1>
<div class="sub">Generated {esc(generated)} &middot; metadata-only (no document contents read) &middot;
expectations derived from holdings &middot; edit <code>{SCHEDULE_FILE}</code> to change deadlines,
earliest tax year, or excluded entities.</div>
<div class="cards">
 <div class="card"><b style="color:#b71c1c">{counts.get('overdue', 0)}</b>overdue K-1s</div>
 <div class="card"><b style="color:#1f4e79">{counts.get('pending', 0)}</b>pending (not yet due)</div>
 <div class="card"><b style="color:#4a148c">{counts.get('review', 0)}</b>needs review</div>
 <div class="card"><b style="color:#8a4b16">{counts.get('locked', 0)}</b>password protected</div>
 <div class="card"><b style="color:#8a6d1a">{counts.get('draft', 0)}</b>draft only, final owed</div>
 <div class="card"><b style="color:#8a6d1a">{counts.get('retag', 0)}</b>needs entity tag</div>
 <div class="card"><b style="color:#455a64">{counts.get('unverified', 0)}</b>fund to verify</div>
 <div class="card"><b style="color:#1b5e20">{counts.get('received', 0) + counts.get('late', 0)}</b>received</div>
</div>"""]

    action = [r for r in recs if r["status"] in CHASE_STATUSES]
    action.sort(key=lambda r: (CHASE_STATUSES.index(r["status"]), -r["tax_year"],
                               r["investment"].lower()))
    if action or undated:
        parts.append("<h2>Action needed</h2><table><tr><th>Status</th><th>Fund</th>"
                     "<th>Chase who</th><th>Entity</th><th>Tax year</th><th>Due</th>"
                     "<th>Detail</th></tr>")
        for r in action:
            lbl, bg, fg = STATUS_META[r["status"]]
            detail = {
                "overdue": f'no K-1 in Canoe &middot; {esc(r["days_late"])} days past due',
                "review": f'{esc(r["document_type"])} &middot; Canoe status: {esc(r["document_status"])}',
                "locked": "K-1 present but encrypted &mdash; request the password",
                "retag": (
                    f'{r["n_untagged"]} untagged K-1 for this fund/year, but '
                    f'{esc(r.get("n_claimants"))} entities are missing it &mdash; the linked '
                    f'document belongs to only <b>one</b> of them. Open it to see which, then '
                    f'assign the entity in Canoe: {esc(r["doc_name"])}'
                    if r.get("ambiguous") else
                    f'K-1 present, no entity assigned: {esc(r["doc_name"])}'),
                "unverified": 'this fund has never issued a K-1 for any year &mdash; confirm '
                              'whether it does, then set <code>track=no</code> if not',
                "draft": 'only a <b>DRAFT</b> K-1 has arrived &mdash; the final is still '
                         f'outstanding: {esc(r["doc_name"])}',
            }[r["status"]]
            if r.get("mistag_check") and r["status"] in ("overdue", "unverified"):
                detail = (f'<b>Check Canoe tagging first.</b> {esc(r["mistag_hint"])} K-1 for '
                          f'this fund/year &mdash; one may be mis-tagged and actually belong '
                          f'here, so verify before chasing the manager.')
            cls = ' class="overdue-row"' if r["status"] == "overdue" else ""
            parts.append(
                f'<tr{cls}><td><span class="pill" style="background:{bg};color:{fg}">{lbl}</span></td>'
                f'<td class="fund">{esc(r["investment"])}</td><td>{esc(r.get("contact") or r["fund_sponsor"])}</td>'
                f'<td>{esc(r["entity"])}</td><td>{r["tax_year"]}</td><td>{esc(r["due"])}</td>'
                f'<td>{detail}</td></tr>')
        lbl, bg, fg = STATUS_META["review"]
        for r in undated:
            parts.append(
                f'<tr><td><span class="pill" style="background:{bg};color:{fg}">{lbl}</span></td>'
                f'<td class="fund">{esc(r["investment"])}</td><td>{esc(r.get("contact") or r["fund_sponsor"])}</td>'
                f'<td>{esc(r["entity"])}</td><td>&mdash;</td><td>&mdash;</td>'
                f'<td>K-1 with no tax year &mdash; set a data date in Canoe: {esc(r["doc_name"])}</td></tr>')
        # K-1s sitting in Canoe's "unknown" bucket: real documents that cannot reach the
        # grid because they have no fund. Canoe already knows they are K-1s, so this is a
        # tagging job in Canoe, not a document to chase from anyone.
        for r in unassigned or []:
            parts.append(
                f'<tr><td><span class="pill" style="background:{bg};color:{fg}">{lbl}</span></td>'
                f'<td class="fund">&mdash;</td><td>&mdash;</td><td>&mdash;</td>'
                f'<td>{esc(r["tax_year"] or "?")}</td><td>&mdash;</td>'
                f'<td>{esc(r["document_type"])} in Canoe with no investment assigned '
                f'&mdash; set the fund in Canoe: {esc(r["doc_name"])}</td></tr>')
        parts.append("</table>")

    cell = {(r["investment"], r["tax_year"]): [] for r in recs}
    for r in recs:
        cell[(r["investment"], r["tax_year"])].append(r)
    parts.append(f'<h2>By fund and tax year ({len(funds)} funds)</h2><table>')
    parts.append("<tr><th>Fund</th>" + "".join(f"<th>TY {y}</th>" for y in years) + "</tr>")
    for sponsor, inv in funds:
        sp = f' <span class="sponsor">({esc(sponsor)})</span>' if sponsor and sponsor != inv else ""
        row = [f'<tr><td class="fund">{esc(inv)}{sp}</td>']
        for y in years:
            rs = cell.get((inv, y)) or []
            if not rs:
                row.append('<td style="color:#d0d7de">&mdash;</td>')
                continue
            # Fund-level roll-up: the worst state across the fund's entities, since
            # that is what needs attention.
            worst = min(rs, key=lambda r: (CHASE_STATUSES.index(r["status"])
                                           if r["status"] in CHASE_STATUSES else 9))
            lbl, bg, fg = STATUS_META[worst["status"]]
            n_ent = len(rs)
            done = sum(1 for r in rs if r["status"] in ("received", "late"))
            tip = esc(f"{done}/{n_ent} entities received; due {worst['due']}")
            txt = lbl if n_ent == 1 else f"{done}/{n_ent}"
            row.append(f'<td><span class="pill" title="{tip}" '
                       f'style="background:{bg};color:{fg}">{txt}</span></td>')
        row.append("</tr>")
        parts.append("".join(row))
    parts.append("</table>")

    with open(path, "w") as f:
        f.write("\n".join(parts))


# --------------------------------------------------------------------------- #
# Digest: K-1s that arrived since the last run
# --------------------------------------------------------------------------- #

def build_digest(backend: str, k1_rows: list[dict], recs: list[dict]) -> tuple[str, list[dict]]:
    """Return (digest_html, new_rows), announcing each K-1 exactly once."""
    state_path = os.path.join(backend, DIGEST_STATE)
    seen: set = set()
    first_run = not os.path.exists(state_path)
    if not first_run:
        try:
            seen = set(json.load(open(state_path)).get("reported_ids", []))
        except (OSError, ValueError):
            first_run = True
    if first_run:
        cutoff = date.today() - timedelta(days=7)
        new = [r for r in k1_rows if r["uploaded"] and r["uploaded"] >= cutoff]
    else:
        new = [r for r in k1_rows if r["doc_id"] and r["doc_id"] not in seen]
    new.sort(key=lambda r: (r["investment"].lower(), r["tax_year"] or 0))

    seen.update(r["doc_id"] for r in k1_rows if r["doc_id"])
    json.dump({"updated": datetime.now(timezone.utc).isoformat(),
               "reported_ids": sorted(seen)}, open(state_path, "w"))

    n_over = sum(1 for r in recs if r["status"] == "overdue")
    n_lock = sum(1 for r in recs if r["status"] == "locked")

    def esc(s):
        return html.escape(str(s))

    body = [f"""<meta charset="utf-8">
<style> body {{ font: 14px -apple-system, "Segoe UI", Helvetica, Arial, sans-serif; color: #24292f; }}
 table {{ border-collapse: collapse; }} th, td {{ border: 1px solid #d8dee4; padding: 4px 10px; text-align: left; }}
 th {{ background: #f6f8fa; }} </style>
<h2>Canoe K-1 digest &mdash; {esc(date.today().isoformat())}</h2>
<p><b>{len(new)}</b> new K-1/K-3 document(s) since the last pull &middot; {n_over} overdue
&middot; {n_lock} password protected.
See the latest <i>K-1 Tracker</i> workbook in the Canoe SharePoint library for the full picture.</p>"""]
    if new:
        body.append("<table><tr><th>Fund</th><th>Entity</th><th>Type</th>"
                    "<th>Tax year</th><th>Status</th><th>Uploaded</th></tr>")
        for r in new:
            ent = r["entity"] if r["entity"] not in IGNORED_ENTITIES else ""
            body.append(f"<tr><td>{esc(r['investment'])}</td><td>{esc(ent)}</td>"
                        f"<td>{esc(r['document_type'])}</td>"
                        f"<td>{esc(r['tax_year'] or 'undated')}</td>"
                        f"<td>{esc(r['document_status'])}</td>"
                        f"<td>{esc(r['uploaded'].isoformat() if r['uploaded'] else '')}</td></tr>")
        body.append("</table>")
    else:
        body.append("<p>No new K-1s this run.</p>")
    return "\n".join(body), new


def email_digest(digest_html: str, n_new: int) -> str:
    """Send the digest if SMTP settings are present. CANOE_K1_DIGEST_TO overrides the
    shared CANOE_DIGEST_TO, so K-1 chasing can go to the tax preparer specifically."""
    to = (os.environ.get("CANOE_K1_DIGEST_TO", "").strip()
          or os.environ.get("CANOE_DIGEST_TO", "").strip())
    user = os.environ.get("CANOE_SMTP_USER", "").strip()
    password = os.environ.get("CANOE_SMTP_PASS", "").strip()
    if not (to and user and password):
        return "email not configured (set CANOE_K1_DIGEST_TO / CANOE_SMTP_USER / CANOE_SMTP_PASS)"
    host = os.environ.get("CANOE_SMTP_HOST", "smtp.office365.com").strip()
    port = int(os.environ.get("CANOE_SMTP_PORT", "587"))
    sender = os.environ.get("CANOE_DIGEST_FROM", user).strip()

    msg = MIMEText(digest_html, "html")
    msg["Subject"] = f"Canoe K-1 digest -- {n_new} new ({date.today().isoformat()})"
    msg["From"] = sender
    msg["To"] = to
    try:
        with smtplib.SMTP(host, port, timeout=60) as s:
            s.starttls()
            s.login(user, password)
            s.sendmail(sender, [a.strip() for a in to.split(",") if a.strip()], msg.as_string())
        return f"emailed to {to}"
    except Exception as exc:                                  # noqa: BLE001
        return f"email FAILED ({exc.__class__.__name__}: {exc})"


# --------------------------------------------------------------------------- #
# Grid rotation
# --------------------------------------------------------------------------- #

def archive_old_grids(outdir: str) -> None:
    """Sweep previous workbooks into Archive/ so the folder shows exactly one."""
    archive_dir = os.path.join(outdir, ARCHIVE)
    for f in sorted(os.listdir(outdir)):
        if not (f.startswith(GRID_PREFIX) and f.endswith(".xlsx")):
            continue
        os.makedirs(archive_dir, exist_ok=True)
        target = os.path.join(archive_dir, f)
        root, ext = os.path.splitext(target)
        n = 2
        while os.path.exists(target):
            target = f"{root}__{n}{ext}"
            n += 1
        os.replace(os.path.join(outdir, f), target)
        print(f"  archived    : {f} -> {ARCHIVE}/{os.path.basename(target)}")


# --------------------------------------------------------------------------- #
# SharePoint destination (--graph)
# --------------------------------------------------------------------------- #

class GraphDest:
    """The tracker's SharePoint destination, reached through Microsoft Graph.

    Same contract as statement_tracker.GraphDest: supply the archive inventory (for
    the grid's hyperlinks) and somewhere to put the outputs. Outputs are BUILT in a
    local staging dir and then uploaded; the metadata cache and digest state stay
    there, because runtime state belongs on local disk, not in the synced library.
    """

    def __init__(self, staging: str):
        import graph_client                       # lazy: local mode needs no Graph creds
        # The inventory callers (dashboard reconcile, canoe_sync --export) hide this
        # folder so its workbooks are not reported as orphaned documents. If the names
        # drift apart that hiding silently stops working, so pin it here.
        assert SUBDIR in graph_client.NON_DOCUMENT_FOLDERS, \
            f"{SUBDIR!r} must be listed in graph_client.NON_DOCUMENT_FOLDERS"
        self.gc = graph_client.GraphClient()
        self.staging = staging
        self.outdir = os.path.join(staging, SUBDIR)
        self.backend = os.path.join(self.outdir, BACKEND)
        # Runtime state, so it stays in staging and never reaches the library.
        self.inv_cache = os.path.join(staging, "inventory_cache.json")
        os.makedirs(self.backend, exist_ok=True)
        sp = config.sharepoint()
        self.label = (f"{sp['hostname']}{sp['site_path']}/{sp['library']}/"
                      f"{self.gc.root_folder}/{SUBDIR}")

    def inventory(self, max_age_sec: int | None = None) -> dict:
        """Listing of the library, used to resolve the grid's hyperlinks.

        Walking ~10k files takes minutes and dominates a run, yet the listing only
        changes when canoe_sync uploads documents -- once a week. `max_age_sec` lets an
        interactive re-publish reuse the previous walk instead of repeating it.

        Off by default, deliberately: the unattended job runs immediately after the sync
        has added documents, which is exactly when a stale listing would produce broken
        links for the newest K-1s. Correct by default, fast on request.
        """
        if max_age_sec:
            try:
                age = time.time() - os.stat(self.inv_cache).st_mtime
                if age < max_age_sec:
                    with open(self.inv_cache) as f:
                        inv = json.load(f)
                    print(f"  archive     : reusing cached listing "
                          f"({len(inv['files'])} files, {age / 60:.0f} min old)")
                    return inv
            except (OSError, ValueError, KeyError):
                pass          # unreadable or stale -- fall through and walk it
        inv = self.gc.list_tree(skip={SUBDIR, st.SUBDIR})
        try:
            with open(self.inv_cache, "w") as f:
                json.dump(inv, f)
        except OSError as exc:                                # noqa: BLE001 -- cache is best-effort
            print(f"  note        : could not cache the listing ({exc})")
        return inv

    def pull_schedule(self) -> bool:
        """Fetch the team-editable schedule into staging. True if we have one.

        The live copy always wins: somebody may have edited a deadline in SharePoint
        since the last run, and a stale local copy would silently revert their work.
        """
        data = self.gc.download(f"{SUBDIR}/{BACKEND}/{SCHEDULE_FILE}")
        path = os.path.join(self.backend, SCHEDULE_FILE)
        if data is None:
            if os.path.exists(path):
                print("  schedule    : not in SharePoint yet; using staged copy")
                return True
            return False
        with open(path, "wb") as f:
            f.write(data)
        print(f"  schedule    : pulled {len(data):,} bytes from SharePoint")
        return True

    def archive_old_grids(self) -> None:
        """Move previous workbooks into Archive/ inside SharePoint. A move (PATCH
        parentReference) keeps each item id, so saved links keep resolving."""
        current = [f for f in self.gc.list_folder(SUBDIR)
                   if not f["is_folder"] and f["name"].startswith(GRID_PREFIX)
                   and f["name"].endswith(".xlsx")]
        if not current:
            return
        import graph_client
        taken = {f["name"] for f in self.gc.list_folder(f"{SUBDIR}/{ARCHIVE}")}
        for f in sorted(current, key=lambda x: x["name"]):
            target = _free_name(f["name"], taken)
            try:
                self.gc.move(f["path"], f"{SUBDIR}/{ARCHIVE}",
                             target if target != f["name"] else None)
            except graph_client.GraphError as exc:
                # 423 = somebody has the workbook open in Excel. Housekeeping must never
                # cost us the update: leave the old grid where it is, publish the new one
                # anyway, and let the next run sweep it up.
                if "423" in str(exc) or "resourceLocked" in str(exc):
                    print(f"  note        : {f['name']} is open in Excel and could not be "
                          f"archived -- leaving it in place, publishing anyway")
                    continue
                raise
            taken.add(target)
            print(f"  archived    : {f['name']} -> {ARCHIVE}/{target}")

    def grid_name(self, base: str) -> str:
        taken = {f["name"] for f in self.gc.list_folder(f"{SUBDIR}/{ARCHIVE}")}
        taken |= {f["name"] for f in self.gc.list_folder(SUBDIR)}
        name = f"{base}.xlsx"
        n = 2
        while name in taken:
            name = f"{base} ({n}).xlsx"
            n += 1
        return name

    def publish(self, names: list[tuple[str, str]]) -> None:
        for rel, _purpose in names:
            local = os.path.join(self.outdir, rel)
            if not os.path.exists(local):
                continue
            with open(local, "rb") as f:
                data = f.read()
            folder = os.path.dirname(rel).replace(os.sep, "/")
            self.gc.upload(data, f"{SUBDIR}/{folder}" if folder else SUBDIR,
                           os.path.basename(rel))
            print(f"  uploaded    : {rel} ({len(data):,} bytes)")


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main() -> None:
    ap = argparse.ArgumentParser(description="Track K-1s received per fund / entity / tax year.")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--dest",
                     help="Archive root as a LOCAL path (a synced SharePoint 'Canoe' folder). "
                          f"Outputs go to <dest>/{SUBDIR}/.")
    src.add_argument("--graph", action="store_true",
                     help="Read the archive from SharePoint via Microsoft Graph and upload "
                          f"outputs to <SP_ROOT_FOLDER>/{SUBDIR}/. Needs no synced folder.")
    ap.add_argument("--refresh", default="auto", choices=["auto", "full"],
                    help="auto = cached + delta by last-modified; full = re-pull everything.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Build everything locally but upload nothing (--graph only).")
    ap.add_argument("--reuse-inventory", nargs="?", const=360, type=int, metavar="MINUTES",
                    help="Reuse the previous SharePoint file listing if it is younger than "
                         "MINUTES (default 360), instead of re-walking the whole library. "
                         "Turns a re-publish from minutes into seconds. NOT used by the "
                         "weekly job: that runs right after new documents are uploaded, "
                         "when a stale listing would mislink the newest K-1s.")
    args = ap.parse_args()

    today = date.today()
    print("K-1 tracker")

    gdest = None
    if args.graph:
        # Unattended job: a missing key or unreachable site should leave one readable
        # line in the log, not a traceback.
        try:
            gdest = GraphDest(config.k1_tracker_dir())
            outdir, backend = gdest.outdir, gdest.backend
            print(f"  destination : {gdest.label} (Graph upload"
                  f"{'; DRY RUN, nothing will be uploaded' if args.dry_run else ''})")
            print(f"  staging     : {outdir}")
            inventory = gdest.inventory(args.reuse_inventory * 60 if args.reuse_inventory else None)
            print(f"  archive     : {len(inventory['files'])} files in the live library")
            have_schedule = gdest.pull_schedule()
        except config.ConfigError as exc:
            sys.exit(f"K-1 tracker: {exc}")
        except Exception as exc:                              # noqa: BLE001 -- incl. GraphError
            sys.exit(f"K-1 tracker: SharePoint unreachable -- "
                     f"{exc.__class__.__name__}: {exc}")
    else:
        dest = os.path.abspath(os.path.expanduser(args.dest))
        outdir = os.path.join(dest, SUBDIR)
        backend = os.path.join(outdir, BACKEND)
        os.makedirs(backend, exist_ok=True)
        print(f"  grid        : {os.path.join(outdir, GRID_PREFIX)} <date>.xlsx")
        print(f"  backend     : {backend}")
        inventory = local_inventory(dest)
        have_schedule = os.path.exists(os.path.join(backend, SCHEDULE_FILE))

    # The manifest is the sync's own record of where each Canoe document was written, so
    # it resolves links that name-matching cannot: a document re-tagged in Canoe after it
    # was filed, or one filed under Unknown Investment before Canoe finished classifying
    # it. Absent or unreadable, links simply fall back to matching on name.
    path_by_doc_id = {}
    try:
        path_by_doc_id = {doc_id: e["dest_path"]
                          for doc_id, e in Manifest(config.manifest_path()).items().items()
                          if e.get("dest_path")}
        print(f"  manifest    : {len(path_by_doc_id)} documents available for exact linking")
    except Exception as exc:                                  # noqa: BLE001 -- best effort
        print(f"  manifest    : unavailable ({exc.__class__.__name__}); "
              f"links will match on document name only")
    links = ArchiveLinks(inventory, web=bool(args.graph), path_by_doc_id=path_by_doc_id)

    # Aliases are read before the rows are built, because they rename entities at source.
    sched_path = os.path.join(backend, SCHEDULE_FILE)
    aliases = load_entity_aliases(sched_path)
    if aliases:
        print(f"  aliases     : {len(aliases)} entity name(s) folded into another "
              f"({', '.join(f'{a} -> {c}' for a, c in sorted(aliases.items()))})")

    docs = load_metadata(os.path.join(backend, CACHE_FILE), args.refresh)
    k1_rows, hold_rows, unassigned = _split(docs, aliases)
    spans = holding_spans(k1_rows, hold_rows)
    print(f"  documents   : {len(docs)} pulled "
          f"({len(K1_TYPE_NAMES)} K-1 types + {len(HOLDING_TYPE_NAMES)} holdings types)")
    print(f"  k-1/k-3     : {len(k1_rows)} allocation rows across "
          f"{len({r['investment'] for r in k1_rows})} funds")
    print(f"  holdings    : {len(spans)} fund-entity pairs "
          f"across {len({i for i, _ in spans})} funds")
    if unassigned:
        print(f"  unassigned  : {len(unassigned)} K-1/K-3 in Canoe with no investment "
              f"-- assign the fund in Canoe and they join the grid")

    sched_before = _digest(sched_path)
    if not have_schedule:
        print(f"  no schedule found -- seeding from holdings (review and edit {SCHEDULE_FILE}!)")
        sched = seed_schedule(k1_rows, hold_rows, spans, sched_path, aliases)
    else:
        sched = sync_new_funds(load_schedule(sched_path), k1_rows, hold_rows, spans,
                               sched_path, aliases)

    recs = reconcile(sched, k1_rows, spans, today, aliases)
    tracked = {s["investment"] for s in sched
               if (s.get("track") or "").strip().lower() in ("yes", "y", "true", "1")}
    undated = sorted((r for r in k1_rows
                      if not r["tax_year"] and r["investment"] in tracked),
                     key=lambda r: r["investment"].lower())
    tally = {k: sum(1 for r in recs if r["status"] == k) for k in STATUS_META}
    print(f"  reconciled  : {len(recs)} fund-entity-years across "
          f"{len({r['investment'] for r in recs})} funds "
          f"| received {tally['received'] + tally['late']} | pending {tally['pending']} "
          f"| overdue {tally['overdue']} | draft-only {tally['draft']} "
          f"| review {tally['review']} | locked {tally['locked']} | retag {tally['retag']} "
          f"| to verify {tally['unverified']} | undated {len(undated)}")

    generated = datetime.now().strftime("%Y-%m-%d %H:%M")
    write_status_csv(os.path.join(backend, "k1_status.csv"), recs)
    write_received_log(os.path.join(backend, "k1_received_log.csv"), k1_rows)
    write_html(os.path.join(backend, "K-1 Tracker.html"), recs, undated, generated, unassigned)

    # Every run writes a brand-new workbook and archives the previous one, same-day
    # reruns included: a reused filename is a reused OneDrive item, and rewriting an
    # item someone has open in Excel wedges its sync.
    base = f"{GRID_PREFIX} {today.isoformat()}"
    if gdest and not args.dry_run:
        gdest.archive_old_grids()
        grid_name = gdest.grid_name(base)
        # Staging keeps only the workbook being uploaded; older ones already live in
        # SharePoint's Archive/ and re-uploading them would undo that sweep.
        for f in os.listdir(outdir):
            if f.startswith(GRID_PREFIX) and f.endswith(".xlsx"):
                os.remove(os.path.join(outdir, f))
    else:
        archive_old_grids(outdir)
        archive_dir = os.path.join(outdir, ARCHIVE)
        prior = sum(1 for f in os.listdir(archive_dir) if f.startswith(base)) \
            if os.path.isdir(archive_dir) else 0
        grid_name = f"{base}.xlsx" if prior == 0 else f"{base} ({prior + 1}).xlsx"
    write_xlsx(os.path.join(outdir, grid_name), recs, links, today)
    print(f"  wrote       : {grid_name} + backend detail (html, csvs)")

    digest_html, new_rows = build_digest(backend, k1_rows, recs)
    with open(os.path.join(backend, "K-1 Digest.html"), "w") as f:
        f.write(digest_html)
    sent = "dry run, not sent" if args.dry_run else email_digest(digest_html, len(new_rows))
    print(f"  digest      : {len(new_rows)} new K-1(s); {sent}")

    if gdest and not args.dry_run:
        # Cache and digest state stay local (runtime state); everything the team reads
        # goes up. The schedule is re-uploaded ONLY if this run changed it, so a
        # concurrent edit in SharePoint is not overwritten by an identical copy.
        publish = [(grid_name, "grid"),
                   (os.path.join(BACKEND, "K-1 Tracker.html"), "dashboard"),
                   (os.path.join(BACKEND, "K-1 Digest.html"), "digest"),
                   (os.path.join(BACKEND, "k1_status.csv"), "status csv"),
                   (os.path.join(BACKEND, "k1_received_log.csv"), "received log")]
        sched_after = _digest(sched_path)
        if sched_after is not None and sched_after != sched_before:
            publish.append((os.path.join(BACKEND, SCHEDULE_FILE), "schedule"))
        gdest.publish(publish)
    elif gdest:
        print("  dry run     : built in staging, nothing uploaded")


if __name__ == "__main__":
    main()
