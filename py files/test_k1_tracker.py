#!/usr/bin/env python3
"""Tests for the K-1 tracker's domain rules and its SharePoint destination.

What is worth pinning down here is everything the statement tracker does NOT do: the
holdings-derived expectation (which fund x entity x tax year combinations are owed a
K-1 at all), the June-30 deadline arithmetic, the extra document states K-1s bring
(password-protected, Canoe-flagged duplicate, never-issued), and the schedule
overrides that let a person trim what the holdings rule inferred.

Fixtures are hand-built metadata records in Canoe's shape, so nothing here needs
credentials or network. The Graph destination runs against the same in-memory fake
library that test_statement_tracker.py uses.
"""

import os
import shutil
import sys
import tempfile
import types
import unittest
from datetime import date

from test_statement_tracker import FakeGraphClient

# The tracker imports graph_client lazily, so the fake must be installed before then.
# Both tracker folders have to be listed: k1_tracker.GraphDest asserts its own folder
# is registered, exactly as the real module does.
_fake_graph = types.ModuleType("graph_client")
_fake_graph.GraphClient = FakeGraphClient
_fake_graph.NON_DOCUMENT_FOLDERS = {"_statement_tracker", "_k1_tracker"}
_fake_graph.GraphError = type("GraphError", (RuntimeError,), {})
sys.modules["graph_client"] = _fake_graph

import config                                                        # noqa: E402
import k1_tracker as k                                               # noqa: E402

TODAY = date(2026, 8, 17)          # after the TY2025 deadline of 2026-06-30


def doc(doc_id, dtype, investment, entity, data_date, *, status="Complete",
        uploaded="2026-03-01", name=None, year=None, original=None):
    """One metadata record in the shape _slim() produces."""
    alloc = {"data_date": data_date, "investment": investment, "investment_id": investment,
             "fund_sponsor": "", "entity": entity, "dataset_type": dtype}
    if year:
        alloc["validated_data"] = year
    return {"id": doc_id, "name": name or f"{investment}-{entity}-{dtype}-{data_date}",
            "original_file_name": original or f"{investment} {dtype}.pdf",
            "document_status": status, "document_type": dtype, "uploaded": uploaded,
            "last_modified": uploaded, "allocations": [alloc]}


def sched_row(investment, **over):
    row = {"investment": investment, "fund_sponsor": "", "contact": "", "track": "yes",
           "first_tax_year": "2022", "due_month_day": "", "exclude_entities": "", "notes": ""}
    row.update(over)
    return row


def reconcile(docs, rows=None, today=TODAY):
    k1_rows, hold_rows, _ = k._split(docs)
    spans = k.holding_spans(k1_rows, hold_rows)
    invs = sorted({r["investment"] for r in k1_rows} | {r["investment"] for r in hold_rows})
    return k.reconcile(rows or [sched_row(i) for i in invs], k1_rows, spans, today)


def status_of(recs, investment, entity, year):
    for r in recs:
        if (r["investment"], r["entity"], r["tax_year"]) == (investment, entity, year):
            return r["status"]
    return None


class TaxYearTests(unittest.TestCase):
    def test_data_date_drives_the_tax_year(self):
        self.assertEqual(k.tax_year({"data_date": "2024-12-31"}), 2024)

    def test_undated_k1_recovers_its_year_from_canoes_extraction(self):
        # 3 of 448 K-1s carry no data_date; validated_data.year agrees with data_date
        # on 428 of 448, so it is a sound fallback. Canoe writes it as "Y2024".
        self.assertEqual(k.tax_year({"data_date": None, "validated_data": "Y2024"}), 2024)
        self.assertEqual(k.tax_year({"data_date": None, "extracted_data": "2023"}), 2023)

    def test_unusable_year_values_do_not_invent_a_year(self):
        for bad in (None, "", "n/a", "Y", "12-31", "1899", "20245"):
            self.assertIsNone(k.tax_year({"data_date": None, "validated_data": bad}), bad)


class DueDateTests(unittest.TestCase):
    def test_deadline_is_the_year_after_the_tax_year(self):
        self.assertEqual(k.due_date(2025, "06-30"), date(2026, 6, 30))

    def test_per_fund_override_is_honoured(self):
        self.assertEqual(k.due_date(2025, "09-15"), date(2026, 9, 15))

    def test_a_malformed_override_falls_back_to_the_default(self):
        # A hand-edited schedule must never crash the unattended job.
        for bad in ("", "bogus", "6/30", None, "13-45"):
            self.assertEqual(k.due_date(2025, bad), date(2026, 6, 30), bad)

    def test_current_tax_year_is_last_calendar_year(self):
        # A K-1 for the year in progress cannot exist, so it is never expected.
        self.assertEqual(k.latest_tax_year(date(2026, 8, 17)), 2025)
        self.assertEqual(k.latest_tax_year(date(2026, 1, 1)), 2025)


class ExpectationFromHoldingsTests(unittest.TestCase):
    """A row exists only where Canoe shows the entity held the fund that year."""

    def test_a_statement_makes_a_k1_expected_for_that_year(self):
        recs = reconcile([doc("s1", "Account Statement", "Acme IV", "Trust A", "2024-06-30")])
        self.assertEqual(status_of(recs, "Acme IV", "Trust A", 2024), "unverified")

    def test_no_holdings_evidence_means_no_row_at_all(self):
        # Nothing is asserted about a fund/entity/year Canoe has never seen.
        recs = reconcile([doc("s1", "Account Statement", "Acme IV", "Trust A", "2024-06-30")])
        self.assertIsNone(status_of(recs, "Acme IV", "Trust B", 2024))
        self.assertIsNone(status_of(recs, "Acme IV", "Trust A", 2022))

    def test_holdings_gaps_are_bridged(self):
        # Canoe's coverage has holes; an entity holding in 2022 and 2025 held throughout.
        recs = reconcile([
            doc("k22", "K-1", "Acme IV", "Trust A", "2022-12-31"),
            doc("s25", "Account Statement", "Acme IV", "Trust A", "2025-03-31"),
        ])
        years = sorted(r["tax_year"] for r in recs if r["entity"] == "Trust A")
        self.assertEqual(years, [2022, 2023, 2024, 2025])

    def test_an_exited_position_stops_being_expected(self):
        # Last evidence in 2023 -> a final K-1 for TY2023, nothing after.
        recs = reconcile([
            doc("s1", "Account Statement", "Acme IV", "Trust A", "2022-12-31"),
            doc("s2", "Account Statement", "Acme IV", "Trust A", "2023-09-30"),
        ])
        self.assertEqual(sorted(r["tax_year"] for r in recs), [2022, 2023])

    def test_current_year_evidence_still_expects_the_last_complete_year(self):
        # A 2026 statement proves the position is live, so TY2025 is owed -- but TY2026
        # is not, because that tax year has not ended.
        recs = reconcile([doc("s1", "Account Statement", "Acme IV", "Trust A", "2026-06-30"),
                          doc("s0", "Account Statement", "Acme IV", "Trust A", "2025-06-30")])
        self.assertEqual(max(r["tax_year"] for r in recs), 2025)

    def test_capital_call_notices_count_as_holdings_evidence(self):
        recs = reconcile([doc("c1", "Capital Call Notice", "Acme IV", "Trust A", "2024-05-01")])
        self.assertEqual(status_of(recs, "Acme IV", "Trust A", 2024), "unverified")

    def test_fund_level_documents_cannot_create_an_entity_row(self):
        # A quarterly report is addressed to all LPs and carries no entity, so it can
        # never establish that a particular entity held the fund.
        self.assertEqual(reconcile([doc("q", "Quarterly Report", "Acme IV", "", "2024-12-31")]), [])

    def test_canoes_unknown_bucket_is_never_tracked(self):
        for junk in ("unknown", "Unknown Investment"):
            self.assertEqual(reconcile([doc("s", "Account Statement", junk, "Trust A",
                                            "2024-12-31")]), [])


class StatusTests(unittest.TestCase):
    def setUp(self):
        # Baseline: the entity holds, and the fund demonstrably issues K-1s.
        self.base = [doc("hold", "Account Statement", "Acme IV", "Trust A", "2025-06-30"),
                     doc("prior", "K-1", "Acme IV", "Trust A", "2024-12-31")]

    def _st(self, *extra, today=TODAY):
        return status_of(reconcile(self.base + list(extra), today=today),
                         "Acme IV", "Trust A", 2025)

    def test_a_clean_k1_is_received(self):
        self.assertEqual(self._st(doc("k", "K-1", "Acme IV", "Trust A", "2025-12-31")), "received")

    def test_a_k1_uploaded_after_the_deadline_is_late(self):
        self.assertEqual(self._st(doc("k", "K-1", "Acme IV", "Trust A", "2025-12-31",
                                      uploaded="2026-08-01")), "late")

    def test_missing_before_the_deadline_is_pending_and_after_is_overdue(self):
        self.assertEqual(self._st(today=date(2026, 6, 29)), "pending")
        self.assertEqual(self._st(today=date(2026, 7, 1)), "overdue")

    def test_a_canoe_flagged_k1_routes_to_review_rather_than_satisfying_the_year(self):
        for flag in ("Awaiting Confirmation", "Anomaly Detected",
                     "Potential Discrepancy", "Configuration Required"):
            self.assertEqual(self._st(doc("k", "K-1", "Acme IV", "Trust A",
                                          "2025-12-31", status=flag)), "review", flag)

    def test_a_password_protected_k1_gets_its_own_state(self):
        # It arrived, so it is not overdue -- but it cannot be read, which is a
        # different action (ask the manager for the password) than a Canoe review.
        self.assertEqual(self._st(doc("k", "K-1", "Acme IV", "Trust A", "2025-12-31",
                                      status="Password Protected")), "locked")

    def test_a_duplicate_never_satisfies_a_year_on_its_own(self):
        self.assertEqual(self._st(doc("k", "K-1", "Acme IV", "Trust A", "2025-12-31",
                                      status="Duplicate")), "overdue")

    def test_a_clean_copy_wins_over_flagged_and_duplicate_siblings(self):
        self.assertEqual(self._st(doc("d", "K-1", "Acme IV", "Trust A", "2025-12-31",
                                      status="Duplicate"),
                                  doc("f", "K-1", "Acme IV", "Trust A", "2025-12-31",
                                      status="Anomaly Detected"),
                                  doc("c", "K-1", "Acme IV", "Trust A", "2025-12-31")), "received")

    def test_an_untagged_k1_for_the_year_asks_for_a_retag(self):
        # The K-1 is in Canoe but nobody assigned the entity: a tagging job, not a
        # missing document, so it must not read as OVERDUE.
        self.assertEqual(self._st(doc("u", "K-1", "Acme IV", "", "2025-12-31")), "retag")

    def test_k3_satisfies_a_year_but_a_k1_is_preferred_for_the_link(self):
        recs = reconcile(self.base + [doc("k3", "K-3", "Acme IV", "Trust A", "2025-12-31"),
                                      doc("k1", "K-1", "Acme IV", "Trust A", "2025-12-31")])
        rec = next(r for r in recs if r["tax_year"] == 2025)
        self.assertEqual(rec["status"], "received")
        self.assertEqual(rec["document_type"], "K-1")


class DraftTests(unittest.TestCase):
    """Canoe types a draft K-1 as "K-1" and marks it Complete, exactly like the real
    thing. Only the manager's filename gives it away."""

    def setUp(self):
        self.base = [doc("hold", "Account Statement", "RMWC DLF II", "JWB Trust", "2026-06-30"),
                     doc("prior", "K-1", "RMWC DLF II", "JWB Trust", "2024-12-31",
                         original="2024 K-1 - RMWC Direct Lending Fund II, LP.pdf")]

    def _st(self, *extra):
        return status_of(reconcile(self.base + list(extra)), "RMWC DLF II", "JWB Trust", 2025)

    def test_draft_markers_are_detected(self):
        for fn in ("2025 DRAFT K-1 - RMWC DLF, LP.pdf",
                   "2022 Estimated K1 RMWC Direct Lending Fund II.pdf",
                   "Preliminary K-1 2025.pdf",
                   "2024 draft k-1.pdf"):
            self.assertTrue(k.is_draft({"original_file_name": fn}), fn)

    def test_final_and_unmarked_filenames_are_not_drafts(self):
        for fn in ("2025 Tax K-1 RMWC Direct Lending Fund II, LP.pdf",
                   "Schedule K1 - FINAL_H00806_123123.pdf",
                   "RMWC DLF II 2021 K-1 - James W.F. Brooks Trust.pdf",
                   "", None):
            self.assertFalse(k.is_draft({"original_file_name": fn}), fn)

    def test_as_amended_in_a_trust_name_is_not_a_draft(self):
        # Every "amend" in this library is the phrase "as amended" inside a trust's
        # legal name. Treating it as a marker would flag dozens of real K-1s.
        self.assertFalse(k.is_draft({"original_file_name":
            "2023 Tax K-1 RMWC DLF II, LP - James W.F. Brooks Trust dated "
            "December 23, 1982, as amended.pdf"}))

    def test_a_draft_alone_does_not_satisfy_the_year(self):
        # The case that prompted this: RMWC Direct Lending II TY2025 has only a draft.
        self.assertEqual(self._st(doc("d", "K-1", "RMWC DLF II", "JWB Trust", "2025-12-31",
                                      original="2025 DRAFT K-1 - RMWC DLF, LP.pdf")), "draft")

    def test_a_final_alongside_a_draft_satisfies_the_year(self):
        recs = reconcile(self.base + [
            doc("d", "K-1", "RMWC DLF II", "JWB Trust", "2025-12-31",
                original="2025 DRAFT K-1 - RMWC DLF, LP.pdf"),
            doc("f", "K-1", "RMWC DLF II", "JWB Trust", "2025-12-31",
                original="2025 Tax K-1 - RMWC DLF, LP.pdf")])
        rec = next(r for r in recs if r["tax_year"] == 2025)
        self.assertEqual(rec["status"], "received")

    def test_the_cell_links_to_the_final_not_the_draft(self):
        recs = reconcile(self.base + [
            doc("d", "K-1", "RMWC DLF II", "JWB Trust", "2025-12-31", uploaded="2026-02-01",
                name="DRAFT-DOC", original="2025 DRAFT K-1.pdf"),
            doc("f", "K-1", "RMWC DLF II", "JWB Trust", "2025-12-31", uploaded="2026-05-01",
                name="FINAL-DOC", original="2025 Tax K-1.pdf")])
        rec = next(r for r in recs if r["tax_year"] == 2025)
        # The draft arrived first; the final must still win the link.
        self.assertEqual(rec["doc_name"], "FINAL-DOC")

    def test_a_draft_is_on_the_chase_list_asking_for_the_final(self):
        self.assertIn("draft", k.ACTION_STATUSES)
        self.assertIn("FINAL", k.CHASE_ACTION["draft"])

    def test_a_draft_is_amber_not_green(self):
        self.assertEqual(k.GRID_STATE["draft"], "attention")
        self.assertNotEqual(k.GRID_STATE["draft"], k.GRID_STATE["received"])


class AttributionTests(unittest.TestCase):
    """Canoe leaves an entity blank when unsure; the manager's filename usually names it."""

    ENTS = ["Pablo & Tiffany Bernal 2016 Trust", "Scott and Amanda Brooks Family Trust",
            "Scott J Brooks and Amanda T Brooks", "Brooks Capital Management, LLC"]

    def _attr(self, filename):
        return k.attribute_untagged({"original_file_name": filename}, self.ENTS)

    def test_ampersand_and_punctuation_differences_still_match(self):
        # Canoe: "Pablo & Tiffany Bernal 2016 Trust"; manager writes it out in full.
        self.assertEqual(self._attr(
            "2024 K-1 Catalyst IOS Fund II LP - James Brooks, Scott Brooks, and Robert L. "
            "Diamond, Trustees of the Pablo and Tiffany Bernal 2016 Trust dated June 3, "
            "2016, and their successors..pdf"), "Pablo & Tiffany Bernal 2016 Trust")

    def test_the_family_trust_is_not_confused_with_the_individuals(self):
        # "Scott and Amanda Brooks Family Trust" and "Scott J Brooks and Amanda T Brooks"
        # are DIFFERENT Canoe entities. Collapsing them is what mis-filed this K-1 in the
        # first place, so the matcher must keep them apart.
        got = self._attr("2024 K-1 Catalyst IOS Fund II LP - James W.F. Brooks, Trustee of "
                         "the Scott and Amanda Brooks Family Trust dated January 25, 2005, "
                         "and his successors..pdf")
        self.assertEqual(got, "Scott and Amanda Brooks Family Trust")

    def test_a_filename_naming_nobody_known_is_left_unattributed(self):
        self.assertIsNone(self._attr("2024 K-1 Some Other Investor LP.pdf"))
        self.assertIsNone(self._attr(""))
        self.assertIsNone(self._attr(None))

    def test_a_filename_naming_two_unrelated_entities_is_left_unattributed(self):
        self.assertIsNone(self._attr(
            "2024 K-1 - Brooks Capital Management, LLC and Scott J Brooks and Amanda T "
            "Brooks combined.pdf"))

    def test_nested_aliases_resolve_to_the_longer_name(self):
        ents = ["Brooks Capital", "Brooks Capital Management, LLC"]
        self.assertEqual(k.attribute_untagged(
            {"original_file_name": "K-1 Brooks Capital Management, LLC.pdf"}, ents),
            "Brooks Capital Management, LLC")

    def test_an_attributed_document_reaches_only_its_own_entity(self):
        # The end-to-end shape of the Catalyst IOS II bug: two entities hold the fund,
        # one untagged K-1 whose filename names the 2016 Trust. Only that row may claim it.
        docs = [
            doc("h1", "Account Statement", "Catalyst IOS II",
                "Pablo & Tiffany Bernal 2016 Trust", "2025-06-30"),
            doc("h2", "Account Statement", "Catalyst IOS II",
                "Scott and Amanda Brooks Family Trust", "2025-06-30"),
            doc("k0", "K-1", "Catalyst IOS II", "Brooks Capital Management, LLC", "2023-12-31"),
            doc("h3", "Account Statement", "Catalyst IOS II",
                "Brooks Capital Management, LLC", "2025-06-30"),
            doc("u", "K-1", "Catalyst IOS II", "--", "2025-12-31",
                original="2025 K-1 Catalyst IOS Fund II LP - James Brooks, Scott Brooks, and "
                         "Robert L. Diamond, Trustees of the Pablo and Tiffany Bernal 2016 "
                         "Trust dated June 3, 2016.pdf"),
        ]
        recs = [r for r in reconcile(docs) if r["tax_year"] == 2025]
        by_ent = {r["entity"]: r for r in recs}
        self.assertEqual(by_ent["Pablo & Tiffany Bernal 2016 Trust"]["status"], "retag")
        self.assertFalse(by_ent["Pablo & Tiffany Bernal 2016 Trust"]["ambiguous"])
        # The S&A Family Trust must NOT be offered the 2016 Trust's document.
        self.assertEqual(by_ent["Scott and Amanda Brooks Family Trust"]["status"], "overdue")


class AmbiguousRetagTests(unittest.TestCase):
    """One untagged K-1 cannot belong to two entities, and the grid must not pretend it
    does -- that is how an entity row ends up linking to another beneficiary's return."""

    def _setup(self, n_entities, n_untagged):
        docs = []
        for i in range(n_entities):
            ent = f"Trust {chr(65+i)}"
            docs.append(doc(f"h{i}", "Account Statement", "Catalyst IOS II", ent, "2025-06-30"))
            docs.append(doc(f"k{i}", "K-1", "Catalyst IOS II", ent, "2023-12-31"))
        for j in range(n_untagged):
            docs.append(doc(f"u{j}", "K-1", "Catalyst IOS II", "--", "2025-12-31",
                            name=f"Catalyst IOS II-K-1-12.31.25-{j}"))
        return [r for r in reconcile(docs) if r["tax_year"] == 2025]

    def test_two_entities_competing_for_one_untagged_k1_are_flagged_ambiguous(self):
        recs = self._setup(2, 1)
        retags = [r for r in recs if r["status"] == "retag"]
        self.assertEqual(len(retags), 2)
        for r in retags:
            self.assertTrue(r["ambiguous"], r["entity"])
            self.assertEqual(r["n_claimants"], 2)
            self.assertEqual(k.cell_label(r), "Tag?")

    def test_a_single_missing_entity_is_not_ambiguous(self):
        recs = self._setup(1, 1)
        retags = [r for r in recs if r["status"] == "retag"]
        self.assertEqual(len(retags), 1)
        self.assertFalse(retags[0]["ambiguous"])
        self.assertEqual(k.cell_label(retags[0]), "Tag")

    def test_the_chase_list_says_ownership_is_unknown(self):
        import openpyxl
        recs = self._setup(2, 1)
        d = tempfile.mkdtemp()
        try:
            p = os.path.join(d, "g.xlsx")
            k.write_xlsx(p, recs, k.ArchiveLinks({"files": [], "folders": []}, web=True), TODAY)
            cs = openpyxl.load_workbook(p)["Chase list"]
            details = " ".join(str(cs.cell(r, 9).value or "") for r in range(5, cs.max_row + 1))
            self.assertIn("only ONE", details)
            self.assertIn("2 entities are missing", details)
        finally:
            shutil.rmtree(d, ignore_errors=True)

    def test_an_unambiguous_retag_keeps_its_plain_wording(self):
        import openpyxl
        recs = self._setup(1, 1)
        d = tempfile.mkdtemp()
        try:
            p = os.path.join(d, "g.xlsx")
            k.write_xlsx(p, recs, k.ArchiveLinks({"files": [], "folders": []}, web=True), TODAY)
            cs = openpyxl.load_workbook(p)["Chase list"]
            details = " ".join(str(cs.cell(r, 9).value or "") for r in range(5, cs.max_row + 1))
            self.assertIn("no entity assigned", details)
            self.assertNotIn("only ONE", details)
        finally:
            shutil.rmtree(d, ignore_errors=True)


class EntityAliasTests(unittest.TestCase):
    """Canoe sometimes files one owner under two names; aliases collapse them."""

    ALIASES = {"scott j brooks and amanda t brooks": "Scott and Amanda Brooks Family Trust"}

    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.path = os.path.join(self.dir, k.SCHEDULE_FILE)

    def tearDown(self):
        shutil.rmtree(self.dir, ignore_errors=True)

    def _docs(self):
        return [
            doc("h", "Account Statement", "Catalyst IOS II",
                "Scott and Amanda Brooks Family Trust", "2025-06-30"),
            doc("k23", "K-1", "Catalyst IOS II",
                "Scott and Amanda Brooks Family Trust", "2023-12-31"),
            # Canoe filed TY2024 under the individuals instead of the trust.
            doc("k24", "K-1", "Catalyst IOS II",
                "Scott J Brooks and Amanda T Brooks", "2024-12-31"),
        ]

    def test_without_an_alias_the_two_names_are_separate_rows(self):
        ents = {r["entity"] for r in reconcile(self._docs())}
        self.assertEqual(len(ents), 2)

    def test_an_alias_folds_them_into_one_row_and_closes_the_gap(self):
        k1, hold, _ = k._split(self._docs(), self.ALIASES)
        spans = k.holding_spans(k1, hold)
        recs = k.reconcile([sched_row("Catalyst IOS II")], k1, spans, TODAY, self.ALIASES)
        self.assertEqual({r["entity"] for r in recs}, {"Scott and Amanda Brooks Family Trust"})
        # The TY2024 K-1 now counts for the trust instead of leaving it overdue. ("late"
        # rather than "received" only because the fixture uploads it after the deadline;
        # what matters is that the year is satisfied.)
        self.assertIn(status_of(recs, "Catalyst IOS II",
                                "Scott and Amanda Brooks Family Trust", 2024),
                      ("received", "late"))

    def test_aliasing_is_case_insensitive(self):
        k1, hold, _ = k._split(self._docs(), {"SCOTT J BROOKS AND AMANDA T BROOKS":
                                              "Scott and Amanda Brooks Family Trust"})
        self.assertEqual({r["entity"] for r in k1 + hold},
                         {"Scott and Amanda Brooks Family Trust"})

    def test_aliases_round_trip_through_the_workbook(self):
        k.write_schedule(self.path, [sched_row("Acme IV")], self.ALIASES)
        self.assertEqual(k.load_entity_aliases(self.path), self.ALIASES)

    def test_a_self_referential_alias_is_ignored(self):
        k.write_schedule(self.path, [sched_row("Acme IV")], {"trust a": "Trust A"})
        self.assertEqual(k.load_entity_aliases(self.path), {})

    def test_appending_a_new_fund_does_not_wipe_hand_entered_aliases(self):
        # write_schedule rewrites the whole workbook, so sync_new_funds must carry the
        # aliases through or a new fund silently destroys them.
        k.write_schedule(self.path, [sched_row("Acme IV")], self.ALIASES)
        docs = self._docs() + [doc("n", "Account Statement", "Zeta VII", "Trust Z", "2025-06-30")]
        nk, nh, _ = k._split(docs, self.ALIASES)
        k.sync_new_funds(k.load_schedule(self.path), nk, nh,
                         k.holding_spans(nk, nh), self.path, self.ALIASES)
        self.assertEqual(k.load_entity_aliases(self.path), self.ALIASES)
        self.assertIn("Zeta VII", {s["investment"] for s in k.load_schedule(self.path)})

    def test_a_workbook_with_no_alias_sheet_reads_as_no_aliases(self):
        self.assertEqual(k.load_entity_aliases(os.path.join(self.dir, "nope.xlsx")), {})

    def test_attribution_matches_an_alias_spelling_and_returns_the_canonical(self):
        got = k.attribute_untagged(
            {"original_file_name": "2024 K-1 - Scott J Brooks and Amanda T Brooks.pdf"},
            ["Scott and Amanda Brooks Family Trust"], self.ALIASES)
        self.assertEqual(got, "Scott and Amanda Brooks Family Trust")


class MistagDetectionTests(unittest.TestCase):
    """Two final K-1s on one entity while a sibling has none is the mis-tag signature."""

    def _docs(self, a_finals, b_finals):
        d = [doc("ha", "Account Statement", "Acme IV", "Trust A", "2025-06-30"),
             doc("hb", "Account Statement", "Acme IV", "Trust B", "2025-06-30"),
             doc("pa", "K-1", "Acme IV", "Trust A", "2023-12-31"),
             doc("pb", "K-1", "Acme IV", "Trust B", "2023-12-31")]
        for i in range(a_finals):
            d.append(doc(f"a{i}", "K-1", "Acme IV", "Trust A", "2025-12-31",
                         name=f"Acme IV-Trust A-K-1-12.31.25-{i}"))
        for i in range(b_finals):
            d.append(doc(f"b{i}", "K-1", "Acme IV", "Trust B", "2025-12-31",
                         name=f"Acme IV-Trust B-K-1-12.31.25-{i}"))
        return [r for r in reconcile(d) if r["tax_year"] == 2025]

    def test_the_missing_entity_is_told_to_check_canoe_not_chase(self):
        recs = self._docs(a_finals=2, b_finals=0)
        b = next(r for r in recs if r["entity"] == "Trust B")
        self.assertEqual(b["status"], "overdue")
        self.assertTrue(b["mistag_check"])
        self.assertIn("Trust A has 2", b["mistag_hint"])
        self.assertEqual(k.cell_label(b), "Check?")

    def test_the_surplus_entity_is_flagged_too(self):
        recs = self._docs(a_finals=2, b_finals=0)
        a = next(r for r in recs if r["entity"] == "Trust A")
        self.assertEqual(a["n_final"], 2)
        self.assertTrue(a["mistag_check"])
        self.assertIn("Trust B has none", a["mistag_hint"])

    def test_one_each_is_not_flagged(self):
        for r in self._docs(a_finals=1, b_finals=1):
            self.assertFalse(r["mistag_check"], r["entity"])

    def test_a_surplus_with_nobody_missing_is_not_flagged(self):
        # An amended K-1 also lands as a second final; with no missing sibling there is
        # nothing to suspect.
        for r in self._docs(a_finals=2, b_finals=1):
            self.assertFalse(r["mistag_check"], r["entity"])

    def test_a_missing_entity_with_no_surplus_anywhere_is_a_normal_chase(self):
        recs = self._docs(a_finals=1, b_finals=0)
        b = next(r for r in recs if r["entity"] == "Trust B")
        self.assertFalse(b["mistag_check"])
        self.assertIsNone(k.cell_label(b))

    def test_a_draft_does_not_count_toward_the_surplus(self):
        d = [doc("ha", "Account Statement", "Acme IV", "Trust A", "2025-06-30"),
             doc("hb", "Account Statement", "Acme IV", "Trust B", "2025-06-30"),
             doc("pa", "K-1", "Acme IV", "Trust A", "2023-12-31"),
             doc("pb", "K-1", "Acme IV", "Trust B", "2023-12-31"),
             doc("a1", "K-1", "Acme IV", "Trust A", "2025-12-31", name="final"),
             doc("a2", "K-1", "Acme IV", "Trust A", "2025-12-31", name="draft",
                 original="2025 DRAFT K-1.pdf")]
        recs = [r for r in reconcile(d) if r["tax_year"] == 2025]
        a = next(r for r in recs if r["entity"] == "Trust A")
        self.assertEqual(a["n_final"], 1, "a draft must not look like a surplus final")
        self.assertFalse(next(r for r in recs if r["entity"] == "Trust B")["mistag_check"])

    def test_the_chase_list_says_check_canoe_before_chasing(self):
        import openpyxl
        recs = self._docs(a_finals=2, b_finals=0)
        d = tempfile.mkdtemp()
        try:
            p = os.path.join(d, "g.xlsx")
            k.write_xlsx(p, recs, k.ArchiveLinks({"files": [], "folders": []}, web=True), TODAY)
            cs = openpyxl.load_workbook(p)["Chase list"]
            actions = " ".join(str(cs.cell(r, 1).value or "") for r in range(5, cs.max_row + 1))
            details = " ".join(str(cs.cell(r, 9).value or "") for r in range(5, cs.max_row + 1))
            self.assertIn("CHECK CANOE TAGGING", actions)
            self.assertNotIn("Chase the manager", actions)
            self.assertIn("may be mis-tagged", details)
        finally:
            shutil.rmtree(d, ignore_errors=True)


class ArchiveLinkTests(unittest.TestCase):
    """Documents are filed under a SANITIZED name, so the lookup must sanitize too."""

    def _links(self, *filenames):
        import statement_tracker as st
        inv = {"files": [{"name": n, "path": f"Fund/{n}", "web_url": f"https://sp/{n}"}
                         for n in filenames],
               "folders": [{"path": "P-E Investments", "name": "P-E Investments",
                            "web_url": "https://sp/folder"}]}
        return st.ArchiveLinks(inv, web=True)

    def test_a_slash_in_the_fund_name_still_resolves_to_the_document(self):
        # Canoe calls it "P/E Investments-...", canoe_sync files it as "P-E Investments-...".
        # Looking up the raw name silently missed and fell back to the folder link.
        links = self._links("P-E Investments-S.J. Brooks, LLC-K-1-12.31.23.pdf")
        url = links.url({"doc_name": "P/E Investments-S.J. Brooks, LLC-K-1-12.31.23",
                         "investment": "P/E Investments"})
        self.assertEqual(url, "https://sp/P-E Investments-S.J. Brooks, LLC-K-1-12.31.23.pdf")

    def test_a_retagged_document_still_resolves_by_id(self):
        # Canoe re-tagging renames the document Canoe reports, but the file already in
        # SharePoint keeps the name it was filed under. The manifest bridges the two.
        import statement_tracker as st
        inv = {"files": [{"name": "Catalyst IOS II-K-1-12.31.25.pdf",
                          "path": "Catalyst IOS II/2025/Tax/Catalyst IOS II-K-1-12.31.25.pdf",
                          "web_url": "https://sp/the-actual-file"}],
               "folders": [{"path": "Catalyst IOS II", "name": "Catalyst IOS II",
                            "web_url": "https://sp/folder"}]}
        rec = {"doc_id": "abc",
               "doc_name": "Catalyst IOS II-Pablo & Tiffany Bernal 2016 Trust-K-1-12.31.25",
               "investment": "Catalyst IOS II"}
        # Without the manifest the name misses and it degrades to the folder.
        self.assertEqual(st.ArchiveLinks(inv, web=True).url(rec), "https://sp/folder")
        # With it, the exact document.
        links = st.ArchiveLinks(inv, web=True, path_by_doc_id={
            "abc": "Catalyst IOS II/2025/Tax/Catalyst IOS II-K-1-12.31.25.pdf"})
        self.assertEqual(links.url(rec), "https://sp/the-actual-file")

    def test_a_document_filed_under_unknown_investment_resolves_by_id(self):
        import statement_tracker as st
        inv = {"files": [{"name": "---K-1-12.31.25.pdf",
                          "path": "Unknown Investment/2025/Tax/---K-1-12.31.25.pdf",
                          "web_url": "https://sp/unknown-filed"}],
               "folders": [{"path": "Catalyst IOS II", "name": "Catalyst IOS II",
                            "web_url": "https://sp/folder"}]}
        links = st.ArchiveLinks(inv, web=True, path_by_doc_id={
            "xyz": "Unknown Investment/2025/Tax/---K-1-12.31.25.pdf"})
        self.assertEqual(links.url({"doc_id": "xyz", "doc_name": "Catalyst IOS II-…-K-1",
                                    "investment": "Catalyst IOS II"}),
                         "https://sp/unknown-filed")

    def test_a_manifest_entry_whose_file_is_gone_falls_through(self):
        import statement_tracker as st
        inv = {"files": [], "folders": [{"path": "Acme IV", "name": "Acme IV",
                                         "web_url": "https://sp/folder"}]}
        links = st.ArchiveLinks(inv, web=True, path_by_doc_id={"a": "Acme IV/gone.pdf"})
        self.assertEqual(links.url({"doc_id": "a", "doc_name": "x", "investment": "Acme IV"}),
                         "https://sp/folder")

    def test_an_ordinary_name_is_unaffected(self):
        links = self._links("Acme IV-Trust A-K-1-12.31.24.pdf")
        self.assertEqual(links.url({"doc_name": "Acme IV-Trust A-K-1-12.31.24",
                                    "investment": "Acme IV"}),
                         "https://sp/Acme IV-Trust A-K-1-12.31.24.pdf")

    def test_an_unknown_document_still_falls_back_to_the_fund_folder(self):
        links = self._links("something else.pdf")
        self.assertEqual(links.url({"doc_name": "not in the library",
                                    "investment": "P/E Investments"}), "https://sp/folder")

    def test_a_blank_doc_name_does_not_match_a_file_called_unfiled(self):
        # _sanitize("") returns "Unfiled"; a blank name must not collide with it.
        links = self._links("Unfiled.pdf")
        self.assertEqual(links.url({"doc_name": "", "investment": "P/E Investments"}),
                         "https://sp/folder")


class UnverifiedFundTests(unittest.TestCase):
    """A fund that has never issued a K-1 is a classification question, not a chase."""

    def test_a_fund_with_no_k1_ever_is_unverified_not_overdue(self):
        recs = reconcile([doc("s", "Account Statement", "Blackstone", "Trust A", "2024-06-30")])
        self.assertEqual(status_of(recs, "Blackstone", "Trust A", 2024), "unverified")

    def test_one_k1_anywhere_makes_later_gaps_a_real_chase(self):
        # Proof the fund issues K-1s comes at FUND level: another entity's K-1 is
        # enough to make this entity's missing year a genuine OVERDUE.
        recs = reconcile([
            doc("s", "Account Statement", "Acme IV", "Trust A", "2025-06-30"),
            doc("k", "K-1", "Acme IV", "Trust B", "2025-12-31"),
            doc("s2", "Account Statement", "Acme IV", "Trust B", "2025-06-30"),
        ])
        self.assertEqual(status_of(recs, "Acme IV", "Trust A", 2025), "overdue")

    def test_an_untagged_k1_still_proves_the_fund_issues_them(self):
        recs = reconcile([doc("s", "Account Statement", "Acme IV", "Trust A", "2024-06-30"),
                          doc("u", "K-1", "Acme IV", "", "2024-12-31")])
        self.assertEqual(status_of(recs, "Acme IV", "Trust A", 2024), "retag")


class ScheduleTests(unittest.TestCase):
    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.path = os.path.join(self.dir, k.SCHEDULE_FILE)
        self.docs = [doc("s", "Account Statement", "Acme IV", "Trust A", "2022-06-30"),
                     doc("s2", "Account Statement", "Acme IV", "Trust A", "2025-06-30"),
                     doc("s3", "Account Statement", "Acme IV", "Trust B", "2025-06-30"),
                     doc("k", "K-1", "Acme IV", "Trust A", "2022-12-31")]

    def tearDown(self):
        shutil.rmtree(self.dir, ignore_errors=True)

    def _seed(self):
        k1_rows, hold_rows, _ = k._split(self.docs)
        spans = k.holding_spans(k1_rows, hold_rows)
        return k.seed_schedule(k1_rows, hold_rows, spans, self.path), k1_rows, spans

    def test_seeding_round_trips_every_column(self):
        seeded, _, _ = self._seed()
        loaded = k.load_schedule(self.path)
        self.assertEqual([s["investment"] for s in loaded], [s["investment"] for s in seeded])
        for col in k.SCHEDULE_HEADER:
            self.assertIn(col, loaded[0], f"{col} lost in the workbook round-trip")

    def test_a_fund_that_never_issued_a_k1_is_flagged_for_a_human(self):
        docs = [doc("s", "Account Statement", "Blackstone", "Trust A", "2024-06-30")]
        k1_rows, hold_rows, _ = k._split(docs)
        seeded = k.seed_schedule(k1_rows, hold_rows,
                                 k.holding_spans(k1_rows, hold_rows), self.path)
        self.assertIn("NO K-1 EVER RECEIVED", seeded[0]["notes"])

    def test_first_tax_year_floors_the_expectation(self):
        _, k1_rows, spans = self._seed()
        recs = k.reconcile([sched_row("Acme IV", first_tax_year="2024")], k1_rows, spans, TODAY)
        self.assertEqual(min(r["tax_year"] for r in recs), 2024)

    def test_seeded_first_tax_year_never_predates_canoes_coverage(self):
        docs = [doc("s", "Account Statement", "Old Fund", "Trust A", "2018-06-30"),
                doc("s2", "Account Statement", "Old Fund", "Trust A", "2025-06-30")]
        k1_rows, hold_rows, _ = k._split(docs)
        seeded = k.seed_schedule(k1_rows, hold_rows,
                                 k.holding_spans(k1_rows, hold_rows), self.path)
        self.assertEqual(seeded[0]["first_tax_year"], str(k.DEFAULT_FIRST_TAX_YEAR))

    def test_exclude_entities_drops_only_that_entity(self):
        _, k1_rows, spans = self._seed()
        recs = k.reconcile([sched_row("Acme IV", exclude_entities="Trust B")],
                           k1_rows, spans, TODAY)
        self.assertEqual({r["entity"] for r in recs}, {"Trust A"})

    def test_exclude_entities_accepts_a_semicolon_list_and_ignores_case(self):
        _, k1_rows, spans = self._seed()
        recs = k.reconcile([sched_row("Acme IV", exclude_entities="trust a ; Trust B")],
                           k1_rows, spans, TODAY)
        self.assertEqual(recs, [])

    def test_track_no_removes_the_fund(self):
        _, k1_rows, spans = self._seed()
        self.assertEqual(k.reconcile([sched_row("Acme IV", track="no")],
                                     k1_rows, spans, TODAY), [])

    def test_pushing_the_deadline_out_turns_overdue_into_pending(self):
        _, k1_rows, spans = self._seed()
        for md, expect in (("06-30", "overdue"), ("12-31", "pending")):
            recs = k.reconcile([sched_row("Acme IV", due_month_day=md)], k1_rows, spans, TODAY)
            self.assertEqual(status_of(recs, "Acme IV", "Trust A", 2025), expect, md)

    def test_contact_reaches_the_chase_list(self):
        _, k1_rows, spans = self._seed()
        recs = k.reconcile([sched_row("Acme IV", contact="ir@acme.example")],
                           k1_rows, spans, TODAY)
        self.assertEqual({r["contact"] for r in recs}, {"ir@acme.example"})

    def test_new_funds_are_appended_and_flagged(self):
        seeded, k1_rows, spans = self._seed()
        extra = self.docs + [doc("n", "Account Statement", "Zeta VII", "Trust A", "2025-06-30")]
        nk, nh, _ = k._split(extra)
        grown = k.sync_new_funds(list(seeded), nk, nh, k.holding_spans(nk, nh), self.path)
        added = next(s for s in grown if s["investment"] == "Zeta VII")
        self.assertTrue(added["notes"].startswith("NEW --"))
        self.assertEqual(len(grown), len(seeded) + 1)

    def test_hand_edits_survive_a_new_fund_being_appended(self):
        seeded, _, _ = self._seed()
        edited = [{**s, "contact": "keep@me.example", "track": "no"} for s in seeded]
        extra = self.docs + [doc("n", "Account Statement", "Zeta VII", "Trust A", "2025-06-30")]
        nk, nh, _ = k._split(extra)
        k.sync_new_funds(edited, nk, nh, k.holding_spans(nk, nh), self.path)
        reloaded = {s["investment"]: s for s in k.load_schedule(self.path)}
        self.assertEqual(reloaded["Acme IV"]["contact"], "keep@me.example")
        self.assertEqual(reloaded["Acme IV"]["track"], "no")


class OutputTests(unittest.TestCase):
    def setUp(self):
        self.dir = tempfile.mkdtemp()
        docs = [doc("h", "Account Statement", "Acme IV", "Trust A", "2025-06-30"),
                doc("k", "K-1", "Acme IV", "Trust A", "2024-12-31"),
                doc("l", "K-1", "Acme IV", "Trust A", "2023-12-31",
                    status="Password Protected"),
                doc("b", "Account Statement", "Blackstone", "Trust B", "2024-06-30")]
        self.recs = reconcile(docs)
        self.k1_rows, _, _ = k._split(docs)

    def tearDown(self):
        shutil.rmtree(self.dir, ignore_errors=True)

    def _grid(self):
        import openpyxl
        path = os.path.join(self.dir, "grid.xlsx")
        k.write_xlsx(path, self.recs, k.ArchiveLinks({"files": [], "folders": []}, web=True), TODAY)
        return openpyxl.load_workbook(path)

    def test_workbook_has_the_four_sheets_and_opens(self):
        self.assertEqual(self._grid().sheetnames,
                         ["By fund", "By entity", "Chase list", "Entity summary"])

    def test_the_two_grids_show_identical_data(self):
        # "By fund" and "By entity" are the same records grouped two ways. If they ever
        # disagree, one of them is lying to whoever prefers that view.
        wb = self._grid()
        def painted(sheet, group_col_is_fund):
            out = {}
            ws = wb[sheet]
            # Locate the header row rather than hard-coding it, so this survives the
            # legend changing length again.
            hdr = next(r for r in range(1, 20)
                       if str(ws.cell(r, 3).value or "").startswith("TY "))
            years = {c: int(str(ws.cell(hdr, c).value)[3:])
                     for c in range(3, ws.max_column + 1) if ws.cell(hdr, c).value}
            group = None
            for r in range(hdr + 1, ws.max_row + 1):
                group = ws.cell(r, 1).value or group
                sub = ws.cell(r, 2).value
                if not sub:
                    continue
                fund, ent = (group, sub) if group_col_is_fund else (sub, group)
                for c, y in years.items():
                    cell = ws.cell(r, c)
                    if cell.fill.patternType:
                        out[(fund, ent, y)] = (cell.fill.fgColor.rgb, cell.value)
            return out
        self.assertEqual(painted("By fund", True), painted("By entity", False))
        self.assertTrue(painted("By fund", True), "grids painted nothing at all")

    def test_every_status_maps_to_a_grid_colour(self):
        # A status with no grid state would raise KeyError mid-write; one with no fill
        # would paint a silently blank cell.
        for status in k.STATUS_META:
            self.assertIn(status, k.GRID_STATE, f"{status} has no grid colour")
        self.assertEqual(set(k.GRID_STATE.values()),
                         {"received", "pending", "outstanding", "attention"})

    def test_the_legend_stays_short(self):
        # It was nine rows and read as noise. Four colours plus "not expected".
        ws = self._grid()["By fund"]
        legend = [ws.cell(r, 3).value for r in range(1, 6)]
        self.assertTrue(all(legend), "a legend row is missing its description")
        self.assertIsNone(ws.cell(6, 2).fill.patternType, "more than five legend rows")

    def test_every_chase_status_has_an_action_and_reaches_the_sheet(self):
        for status in k.CHASE_STATUSES:
            self.assertIn(status, k.CHASE_ACTION, f"{status} has no Chase list action")
        cs = self._grid()["Chase list"]
        listed = {cs.cell(r, 2).value for r in range(5, cs.max_row + 1)}
        expected = {k.STATUS_META[r["status"]][0] for r in self.recs
                    if r["status"] in k.CHASE_STATUSES}
        self.assertEqual(listed - {None}, expected)

    def test_chase_list_separates_chasing_from_confirming(self):
        # The whole point of the Action column: a fund that has filed before is a
        # manager to chase; one that never has may not issue K-1s at all.
        self.assertEqual(k.CHASE_ACTION["overdue"], "Chase the manager")
        self.assertNotEqual(k.CHASE_ACTION["unverified"], k.CHASE_ACTION["overdue"])
        cs = self._grid()["Chase list"]
        actions = {cs.cell(r, 1).value for r in range(5, cs.max_row + 1)} - {None}
        self.assertIn("Confirm this fund issues a K-1", actions)

    def test_entity_sheet_totals_reconcile_with_the_records(self):
        ws = self._grid()["Entity summary"]
        years = sorted({r["tax_year"] for r in self.recs})
        col_expected = 2 + len(years)
        rows = {ws.cell(r, 1).value: r for r in range(5, ws.max_row + 1) if ws.cell(r, 1).value}
        self.assertEqual(set(rows), {r["entity"] for r in self.recs})
        for entity, row in rows.items():
            mine = [r for r in self.recs if r["entity"] == entity]
            self.assertEqual(ws.cell(row, col_expected).value, len(mine), entity)
            got = sum(1 for r in mine if r["status"] in ("received", "late"))
            self.assertEqual(ws.cell(row, col_expected + 1).value, got, entity)

    def test_unassigned_k1s_are_reported_not_dropped(self):
        # Canoe's "unknown" bucket holds real K-1s; they cannot reach the grid, but
        # silently discarding them would hide documents somebody is waiting on.
        docs = [doc("u", "K-1", "unknown", "--", "2025-12-31")]
        k1_rows, hold_rows, unassigned = k._split(docs)
        self.assertEqual((k1_rows, hold_rows), ([], []))
        self.assertEqual(len(unassigned), 1)
        self.assertEqual(unassigned[0]["tax_year"], 2025)
        path = os.path.join(self.dir, "u.html")
        k.write_html(path, self.recs, [], "now", unassigned)
        with open(path) as f:
            self.assertIn("no investment assigned", f.read())

    def test_status_csv_survives_records_with_differing_keys(self):
        # Regression: n_claimants was set only on ambiguous rows, so a DictWriter taking
        # its columns from the first record aborted the run on a later one -- at the very
        # last step, after all the API work was done.
        import csv as _csv
        recs = [dict(self.recs[0]), {**self.recs[0], "surprise_field": "x"}]
        path = os.path.join(self.dir, "s.csv")
        k.write_status_csv(path, recs)
        with open(path) as f:
            header = next(_csv.reader(f))
        self.assertIn("surprise_field", header)

    def test_every_record_has_the_same_keys(self):
        # The root cause, asserted directly: a heterogeneous record type is the bug.
        shapes = {frozenset(r) for r in self.recs}
        self.assertEqual(len(shapes), 1, "reconcile() emitted records with differing keys")

    def test_html_and_csvs_are_written(self):
        k.write_html(os.path.join(self.dir, "t.html"), self.recs, [], "2026-08-17 09:00")
        k.write_status_csv(os.path.join(self.dir, "s.csv"), self.recs)
        k.write_received_log(os.path.join(self.dir, "r.csv"), self.k1_rows)
        for f in ("t.html", "s.csv", "r.csv"):
            self.assertGreater(os.path.getsize(os.path.join(self.dir, f)), 0, f)

    def test_html_escapes_a_hostile_fund_name(self):
        recs = [{**self.recs[0], "investment": '<script>alert(1)</script>', "status": "overdue"}]
        path = os.path.join(self.dir, "x.html")
        k.write_html(path, recs, [], "now")
        with open(path) as f:
            body = f.read()
        self.assertNotIn("<script>alert(1)</script>", body)
        self.assertIn("&lt;script&gt;", body)

    def test_digest_announces_each_k1_once(self):
        first, new1 = k.build_digest(self.dir, self.k1_rows, self.recs)
        self.assertIn("K-1 digest", first)
        _, new2 = k.build_digest(self.dir, self.k1_rows, self.recs)
        self.assertEqual(new2, [], "a second run must not re-announce the same K-1s")


class GraphDestTests(unittest.TestCase):
    """The --graph seam: staging, schedule pull, archive sweep, publish."""

    def setUp(self):
        self.staging = tempfile.mkdtemp()
        self.dest = k.GraphDest(self.staging)
        self.gc = self.dest.gc

    def tearDown(self):
        shutil.rmtree(self.staging, ignore_errors=True)

    def test_output_folder_is_registered_as_non_document(self):
        # canoe_sync --export and the dashboard hide these folders; if the name drifts
        # the workbooks start showing up as orphaned documents.
        import graph_client
        self.assertIn(k.SUBDIR, graph_client.NON_DOCUMENT_FOLDERS)

    def test_inventory_is_rewalked_by_default(self):
        # The unattended job must never reuse a listing: it runs right after new
        # documents land, which is when a stale one would mislink the newest K-1s.
        self.gc.lib = {"Acme IV/k1.pdf": b"x"}
        self.dest.inventory()
        self.gc.lib["Acme IV/new.pdf"] = b"y"
        self.assertEqual(len(self.dest.inventory()["files"]), 2)

    def test_a_fresh_cached_inventory_is_reused(self):
        self.gc.lib = {"Acme IV/k1.pdf": b"x"}
        self.dest.inventory()                       # populates the cache
        self.gc.lib["Acme IV/new.pdf"] = b"y"       # library moves on
        self.assertEqual(len(self.dest.inventory(max_age_sec=3600)["files"]), 1)

    def test_an_expired_cached_inventory_is_rewalked(self):
        self.gc.lib = {"Acme IV/k1.pdf": b"x"}
        self.dest.inventory()
        os.utime(self.dest.inv_cache, (0, 0))       # pretend it is ancient
        self.gc.lib["Acme IV/new.pdf"] = b"y"
        self.assertEqual(len(self.dest.inventory(max_age_sec=60)["files"]), 2)

    def test_a_missing_cache_falls_back_to_walking(self):
        self.gc.lib = {"Acme IV/k1.pdf": b"x"}
        self.assertEqual(len(self.dest.inventory(max_age_sec=3600)["files"]), 1)

    def test_the_inventory_cache_stays_out_of_the_library(self):
        self.gc.lib = {"Acme IV/k1.pdf": b"x"}
        self.dest.inventory()
        self.assertTrue(os.path.exists(self.dest.inv_cache))
        self.dest.publish([("grid.xlsx", "grid")])
        self.assertNotIn("inventory_cache.json", " ".join(self.gc.lib))

    def test_inventory_skips_both_trackers_own_folders(self):
        self.gc.lib = {"Acme IV/2025/k1.pdf": b"x",
                       f"{k.SUBDIR}/K-1 Tracker 2026-08-10.xlsx": b"x",
                       "_statement_tracker/Statement Tracker 2026-08-10.xlsx": b"x"}
        self.assertEqual([f["path"] for f in self.dest.inventory()["files"]],
                         ["Acme IV/2025/k1.pdf"])

    def test_schedule_pull_prefers_the_live_copy(self):
        # Somebody may have edited a deadline in SharePoint since the last run.
        self.gc.lib[f"{k.SUBDIR}/{k.BACKEND}/{k.SCHEDULE_FILE}"] = b"live-copy"
        staged = os.path.join(self.dest.backend, k.SCHEDULE_FILE)
        with open(staged, "wb") as f:
            f.write(b"stale-local")
        self.assertTrue(self.dest.pull_schedule())
        with open(staged, "rb") as f:
            self.assertEqual(f.read(), b"live-copy")

    def test_schedule_pull_reports_absence_on_a_first_run(self):
        self.assertFalse(self.dest.pull_schedule())

    def test_old_grids_are_moved_not_recreated(self):
        # A move keeps the item id, so links people saved keep resolving.
        self.gc.lib[f"{k.SUBDIR}/K-1 Tracker 2026-08-10.xlsx"] = b"old"
        self.dest.archive_old_grids()
        self.assertIn(f"{k.SUBDIR}/{k.ARCHIVE}/K-1 Tracker 2026-08-10.xlsx", self.gc.lib)
        self.assertNotIn(f"{k.SUBDIR}/K-1 Tracker 2026-08-10.xlsx", self.gc.lib)
        self.assertEqual(len(self.gc.moves), 1)

    def test_archiving_a_same_named_grid_does_not_overwrite(self):
        self.gc.lib[f"{k.SUBDIR}/{k.ARCHIVE}/K-1 Tracker 2026-08-10.xlsx"] = b"first"
        self.gc.lib[f"{k.SUBDIR}/K-1 Tracker 2026-08-10.xlsx"] = b"second"
        self.dest.archive_old_grids()
        self.assertIn(f"{k.SUBDIR}/{k.ARCHIVE}/K-1 Tracker 2026-08-10__2.xlsx", self.gc.lib)
        self.assertEqual(self.gc.lib[f"{k.SUBDIR}/{k.ARCHIVE}/K-1 Tracker 2026-08-10.xlsx"],
                         b"first")

    def test_a_workbook_open_in_excel_does_not_abort_the_run(self):
        # A locked file cannot be moved (HTTP 423). Housekeeping must degrade, not fail:
        # somebody reviewing the grid would otherwise block the whole scheduled run.
        import graph_client
        self.gc.lib[f"{k.SUBDIR}/K-1 Tracker 2026-08-10.xlsx"] = b"open in excel"

        def locked(*a, **kw):
            raise graph_client.GraphError("Move failed: 423 {'code':'resourceLocked'}")
        self.gc.move = locked
        self.dest.archive_old_grids()                     # must not raise
        self.assertIn(f"{k.SUBDIR}/K-1 Tracker 2026-08-10.xlsx", self.gc.lib)

    def test_a_genuine_move_failure_still_raises(self):
        import graph_client
        self.gc.lib[f"{k.SUBDIR}/K-1 Tracker 2026-08-10.xlsx"] = b"x"

        def broken(*a, **kw):
            raise graph_client.GraphError("Move failed: 500 server exploded")
        self.gc.move = broken
        with self.assertRaises(graph_client.GraphError):
            self.dest.archive_old_grids()

    def test_grid_name_avoids_names_taken_anywhere(self):
        base = "K-1 Tracker 2026-08-17"
        self.gc.lib[f"{k.SUBDIR}/{base}.xlsx"] = b"x"
        self.gc.lib[f"{k.SUBDIR}/{k.ARCHIVE}/{base} (2).xlsx"] = b"x"
        self.assertEqual(self.dest.grid_name(base), f"{base} (3).xlsx")

    def test_publish_uploads_staged_files_and_skips_absent_ones(self):
        with open(os.path.join(self.dest.outdir, "grid.xlsx"), "wb") as f:
            f.write(b"g")
        with open(os.path.join(self.dest.backend, "k1_status.csv"), "wb") as f:
            f.write(b"c")
        self.dest.publish([("grid.xlsx", "grid"),
                           (os.path.join(k.BACKEND, "k1_status.csv"), "csv"),
                           (os.path.join(k.BACKEND, "not-there.csv"), "missing")])
        self.assertEqual(sorted(self.gc.lib),
                         [f"{k.SUBDIR}/{k.BACKEND}/k1_status.csv", f"{k.SUBDIR}/grid.xlsx"])

    def test_cache_and_digest_state_never_leave_the_staging_dir(self):
        # Runtime state belongs on local disk, not in the shared library.
        for name in (k.CACHE_FILE, k.DIGEST_STATE):
            with open(os.path.join(self.dest.backend, name), "wb") as f:
                f.write(b"state")
        self.dest.publish([(os.path.join(k.BACKEND, "k1_status.csv"), "csv")])
        self.assertEqual(self.gc.lib, {})


if __name__ == "__main__":
    unittest.main()
