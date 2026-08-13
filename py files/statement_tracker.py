#!/usr/bin/env python3
"""
statement_tracker.py -- Track which manager statements have arrived, per period.

Architecture A ("Statement Tracker") v1: metadata-and-rules only. The tool reads
Canoe's structured document metadata (fund, sponsor, data date, type, status) via
GET /v1/documents/data -- it never opens a document body, so it carries no GenAI
dependency and no new data-handling surface.

How it works
------------
1. Pull metadata for every statement-type document -- Account Statement,
   Quarterly Report, Financials, etc. -- across ALL Canoe categories (Canoe
   scatters these types over Capital Activity, Investment Reporting, and
   Financial Statements & Performance). Full pull the first time, incremental
   by last-modified after that; cached in a JSON file beside the outputs.
2. Load -- or, on first run, auto-seed from history -- an editable schedule:
   one row per fund with its expected frequency (monthly/quarterly/annual) and
   a grace period in days after period end.
3. Reconcile: for each tracked fund and each period since tracking start,
   decide Received / Pending / OVERDUE / Review.
4. Write outputs beside the archive (SharePoint-synced, team-visible):
   - _statement_tracker/Statement Tracker.xlsx  -- THE team grid: green = received,
     red = not; one sheet per cadence. The only file in the folder.
   - _statement_tracker/backend/               -- supporting detail: HTML status
     dashboard, status/received CSVs, the editable schedule workbook
     (statement_schedule.xlsx), and the metadata cache.

Canoe review flags never auto-confirm a period: a document whose status is not
Complete routes to the Review column instead of silently satisfying the period.

Usage:
  python statement_tracker.py --dest "$CANOE_ARCHIVE_DIR"            # weekly
  python statement_tracker.py --dest "$CANOE_ARCHIVE_DIR" --refresh full
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import os
import re
import smtplib
import sys
import time
import urllib.parse
from datetime import date, datetime, timedelta, timezone
from email.mime.text import MIMEText

import openpyxl
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

import canoe_auth

DATA_URL = "https://api.canoesoftware.com/v1/documents/data"
PAGE_LIMIT = 100
MAX_RETRIES = 5

# Document types that count as "the statement arrived" for a period. The pull
# filters on document_type across ALL Canoe categories -- the same "Account
# Statement" type shows up under Capital Activity, Investment Reporting, and
# Financial Statements & Performance depending on the document, so filtering
# by category silently drops real statements.
DEFAULT_STATEMENT_TYPE_NAMES = [
    "Account Statement",
    "Capital Account Statement",
    "Monthly Report",
    "Quarterly Report",
    "Annual Report",
    "Financials",
]
DEFAULT_STATEMENT_TYPES = {t.lower() for t in DEFAULT_STATEMENT_TYPE_NAMES}

# Canoe document statuses that must be looked at by a person before the
# period they cover can be trusted.
REVIEW_STATUSES = {"awaiting confirmation", "anomaly detected", "potential discrepancy"}

# Default days after period end before a missing statement turns OVERDUE.
DEFAULT_GRACE = {"monthly": 45, "quarterly": 90, "annual": 180}
# Q4 statements routinely arrive with the audit; give them extra slack.
Q4_EXTRA_DAYS = 30

SUBDIR = "_statement_tracker"
BACKEND = "backend"
ARCHIVE = "Archive"
SCHEDULE_FILE = "statement_schedule.xlsx"
CACHE_FILE = "statement_metadata_cache.json"
# Each run writes a NEW dated workbook ("Statement Tracker 2026-08-11.xlsx")
# and sweeps older ones into Archive/. A fresh file is a fresh OneDrive item,
# so an Excel session holding last week's grid open can never block the
# weekly update from reaching SharePoint (in-place rewrites of an open .xlsx
# wedge OneDrive's Office-file sync).
GRID_PREFIX = "Statement Tracker"


# --------------------------------------------------------------------------- #
# Metadata pull (cached + incremental)
# --------------------------------------------------------------------------- #

def _fetch_page(page: int, extra: dict) -> requests.Response:
    # NB: do NOT pass `fields` here -- when it is supplied this endpoint returns
    # empty `allocations` (losing fund/data-date) and ignores page/limit.
    params = {"page": page, "limit": PAGE_LIMIT, **extra}
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(DATA_URL, headers=canoe_auth.get_auth_headers(),
                                params=params, timeout=600)
        except requests.exceptions.RequestException as exc:
            wait = 15 * attempt
            print(f"    network error ({exc.__class__.__name__}); retry in {wait}s ({attempt}/{MAX_RETRIES})...")
            time.sleep(wait)
            continue
        if resp.status_code in (429, 500, 502, 503, 504):
            wait = 30 * attempt
            print(f"    HTTP {resp.status_code}; retry in {wait}s ({attempt}/{MAX_RETRIES})...")
            time.sleep(wait)
            continue
        resp.raise_for_status()
        return resp
    raise RuntimeError(f"Metadata page {page} failed after {MAX_RETRIES} attempts.")


def _slim(doc: dict) -> dict:
    """Keep only the fields the tracker needs -- the cache stays small."""
    # Canoe sometimes nests allocation entries one level deep (a list inside
    # the list); flatten and keep only dicts.
    raw = doc.get("allocations") or []
    flat = []
    for a in raw:
        if isinstance(a, list):
            flat.extend(x for x in a if isinstance(x, dict))
        elif isinstance(a, dict):
            flat.append(a)
    allocs = []
    for a in flat:
        allocs.append({
            "data_date": a.get("data_date"),
            "investment": a.get("investment"),
            "investment_id": a.get("investment_id"),
            "fund_sponsor": a.get("fund_sponsor"),
            "entity": a.get("entity"),
            "frequency": a.get("frequency"),
            "dataset_type": a.get("dataset_type"),
        })
    return {
        "id": doc.get("id"),
        "name": doc.get("name"),
        "document_status": doc.get("document_status"),
        "document_type": doc.get("document_type"),
        "reporting_frequency": doc.get("reporting_frequency"),
        "uploaded": doc.get("uploaded"),
        "last_modified": doc.get("last_modified"),
        "allocations": allocs,
    }


def _pull(extra: dict, label: str) -> list[dict]:
    # Pagination is treated defensively (the endpoint has quirks): dedupe by id
    # and stop as soon as a page contributes nothing new.
    by_id: dict[str, dict] = {}
    page = 1
    t0 = time.time()
    while page <= 200:
        resp = _fetch_page(page, extra)
        batch = resp.json()
        if not isinstance(batch, list) or not batch:
            break
        before = len(by_id)
        for d in batch:
            slim = _slim(d)
            if slim.get("id"):
                by_id[slim["id"]] = slim
        total_pages = resp.headers.get("total_pages")
        print(f"  {label}: page {page}{'/' + total_pages if total_pages else ''}"
              f" (+{len(by_id) - before} new, total {len(by_id)}, {time.time()-t0:.0f}s)")
        if len(by_id) == before:
            break
        if total_pages and page >= int(total_pages):
            break
        page += 1
        remaining = resp.headers.get("X-RateLimit-Remaining")
        if remaining is not None and int(remaining) <= 3:
            print("    near rate limit; pausing 60s...")
            time.sleep(60)
    return list(by_id.values())


def load_metadata(cache_path: str, refresh: str, type_names: list[str]) -> list[dict]:
    """Return the slim doc list for the given document types (all categories),
    from cache + incremental refresh (or full pull)."""
    types_param = ",".join(sorted(type_names))
    cache = None
    if refresh != "full" and os.path.exists(cache_path):
        try:
            cache = json.load(open(cache_path))
        except (OSError, ValueError):
            cache = None
    # A schedule edit can add doc_types the cache has never pulled -- re-pull.
    if cache is not None and set(cache.get("types", [])) != set(t.lower() for t in type_names):
        print("  document-type set changed -- full re-pull")
        cache = None
    # Deltas can't notice a doc re-typed OUT of the statement set (the type
    # filter stops returning it), so re-baseline in full once a month.
    if cache is not None:
        full_at = cache.get("full_pulled_at")
        if not full_at or (datetime.now(timezone.utc)
                           - datetime.fromisoformat(full_at)).days >= 30:
            print("  monthly re-baseline -- full re-pull")
            cache = None

    if cache is None:
        print("  full metadata pull (first run or --refresh full) -- this can take a while...")
        docs = _pull({"document_type": types_param}, "full pull")
        by_id = {d["id"]: d for d in docs if d.get("id")}
        full_pulled_at = datetime.now(timezone.utc).isoformat()
    else:
        by_id = {d["id"]: d for d in cache.get("docs", []) if d.get("id")}
        # Re-pull anything modified since just before the last run, so re-tags
        # (fund/date corrections done inside Canoe) are picked up too.
        since = (datetime.fromisoformat(cache["pulled_at"]) - timedelta(days=2)).strftime("%Y-%m-%d")
        fresh = _pull({"document_type": types_param,
                       "last_modified_time_start": since}, f"delta since {since}")
        for d in fresh:
            if d.get("id"):
                by_id[d["id"]] = d
        print(f"  cache: {len(by_id)} docs after merging {len(fresh)} modified")
        full_pulled_at = cache.get("full_pulled_at") or datetime.now(timezone.utc).isoformat()

    json.dump({"pulled_at": datetime.now(timezone.utc).isoformat(),
               "full_pulled_at": full_pulled_at,
               "types": sorted(t.lower() for t in type_names),
               "docs": list(by_id.values())}, open(cache_path, "w"))
    return list(by_id.values())


# --------------------------------------------------------------------------- #
# Statement rows
# --------------------------------------------------------------------------- #

def parse_date(s: str | None) -> date | None:
    if not s:
        return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(s))
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def merrill_stems(dest: str) -> frozenset:
    """Lowercase file stems (dedup __N suffix stripped) of everything routed to
    Merrill/. canoe_route.py puts custodian statements there after verifying
    the PDF text locally; the tracker trusts that folder as the exclusion list.
    """
    stems = set()
    merrill_dir = os.path.join(dest, "Merrill")
    for root, _dirs, files in os.walk(merrill_dir):
        for f in files:
            if not f.startswith("."):
                stems.add(re.sub(r"__\d+$", "", os.path.splitext(f)[0]).lower())
    return frozenset(stems)


def statement_rows(docs: list[dict], types_by_fund: dict,
                   excluded_stems: frozenset = frozenset()) -> list[dict]:
    """Flatten docs -> one row per (doc, allocation) that is a statement type.

    Two kinds of custodian statements are excluded -- managers send their own
    statements separately, so custodian copies must not satisfy a fund period:
    - documents whose allocations span more than one investment (consolidated
      brokerage statements Canoe maps to every fund they mention), and
    - documents whose file has been routed to the archive's Merrill/ folder
      (single-fund custodian statements, content-verified by canoe_route.py).
    """
    excluded_span = excluded_merrill = 0
    rows = []
    for d in docs:
        dtype = (d.get("document_type") or "").strip()
        span = {a.get("investment_id") for a in d.get("allocations") or []
                if isinstance(a, dict) and a.get("investment_id")}
        if len(span) > 1:
            excluded_span += 1
            continue
        if (d.get("name") or "").strip().lower() in excluded_stems:
            excluded_merrill += 1
            continue
        for a in d.get("allocations") or []:
            inv = (a.get("investment") or "").strip()
            if not inv or inv.lower() in IGNORED_INVESTMENTS:
                continue
            allowed = types_by_fund.get(inv, DEFAULT_STATEMENT_TYPES)
            ds_type = (a.get("dataset_type") or "").strip()
            if dtype.lower() not in allowed and ds_type.lower() not in allowed:
                continue
            rows.append({
                "investment": inv,
                "investment_id": a.get("investment_id") or "",
                "fund_sponsor": (a.get("fund_sponsor") or "").strip(),
                "entity": (a.get("entity") or "").strip(),
                "frequency_canoe": (a.get("frequency") or d.get("reporting_frequency") or "").strip(),
                "data_date": parse_date(a.get("data_date")),
                "uploaded": parse_date(d.get("uploaded")),
                "document_type": dtype or ds_type,
                "document_status": (d.get("document_status") or "").strip(),
                "doc_id": d.get("id"),
                "doc_name": d.get("name") or "",
            })
    if excluded_span or excluded_merrill:
        print(f"  excluded    : {excluded_span} consolidated (multi-investment) + "
              f"{excluded_merrill} Merrill-routed custodian statements")
    return rows


# --------------------------------------------------------------------------- #
# Periods
# --------------------------------------------------------------------------- #

def month_end(y: int, m: int) -> date:
    return (date(y + (m == 12), (m % 12) + 1, 1) - timedelta(days=1))


def period_of(d: date, freq: str) -> str:
    if freq == "monthly":
        return f"{d.year}-{d.month:02d}"
    if freq == "quarterly":
        return f"{d.year}-Q{(d.month - 1) // 3 + 1}"
    return str(d.year)


def period_end(p: str) -> date:
    if "-Q" in p:
        y, q = p.split("-Q")
        return month_end(int(y), int(q) * 3)
    if "-" in p:
        y, m = p.split("-")
        return month_end(int(y), int(m))
    return date(int(p), 12, 31)


def period_label(p: str) -> str:
    if "-Q" in p:
        return p.replace("-", " ")
    if "-" in p:
        y, m = p.split("-")
        return f"{['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][int(m)-1]} {y}"
    return f"FY {p}"


def periods_between(start: date, end: date, freq: str) -> list[str]:
    """All periods whose period-end falls in [start, end]."""
    out, seen = [], set()
    cur = date(start.year, start.month, 1)
    while cur <= end:
        p = period_of(month_end(cur.year, cur.month), freq)
        if p not in seen and start <= period_end(p) <= end:
            seen.add(p)
            out.append(p)
        cur = date(cur.year + (cur.month == 12), (cur.month % 12) + 1, 1)
    return out


# --------------------------------------------------------------------------- #
# Schedule (editable config, auto-seeded)
# --------------------------------------------------------------------------- #

SCHEDULE_HEADER = ["investment", "fund_sponsor", "frequency", "grace_days",
                   "track", "start_date", "doc_types", "notes"]

def infer_frequency(dates: list[date], canoe_freq: str) -> tuple[str, str]:
    """Return (frequency, source).

    Two signals: the filing history (12-month window anchored at the fund's
    LAST data date, so a fund that has fallen behind still shows its true
    cadence) and Canoe's own frequency field. When they disagree, the LESS
    frequent one wins: custodian notice feeds (e.g. Merrill) make quarterly
    funds look monthly, and Canoe's field sometimes says Monthly for quarterly
    PE funds -- in both cases the sparser cadence is the honest one. Bias is
    toward fewer false "missing statement" alarms; genuinely monthly funds
    that lose the coin toss are one edit away in the schedule.
    """
    cf = canoe_freq.lower()
    canoe = ("monthly" if "month" in cf else
             "quarterly" if "quarter" in cf else
             "annual" if ("annual" in cf or "year" in cf) else None)

    hist = None
    if dates:
        anchor = max(dates)
        horizon = anchor - timedelta(days=365)
        months = {(d.year, d.month) for d in dates if horizon <= d <= anchor}
        n = len(months)
        if n >= 9:
            hist = "monthly"
        elif n >= 3:
            qe = sum(1 for (_, m) in months if m in (3, 6, 9, 12))
            hist = "quarterly" if qe >= max(2, int(0.7 * n)) else "monthly"
        elif n >= 1 and all(m == 12 for (_, m) in months):
            hist = "annual"

    if hist and canoe and hist != canoe:
        rank = {"monthly": 0, "quarterly": 1, "annual": 2}
        pick = max(hist, canoe, key=lambda f: rank[f])
        return pick, f"history says {hist}, canoe says {canoe}; less frequent wins"
    if hist:
        return hist, "history"
    if canoe:
        return canoe, "canoe"
    if dates:
        return "quarterly", "sparse history"
    return "none", "no statements seen"


def seed_schedule(rows: list[dict], path: str) -> list[dict]:
    """First run: derive a schedule row per fund from observed history."""
    by_fund: dict[str, list[dict]] = {}
    for r in rows:
        by_fund.setdefault(r["investment"], []).append(r)

    sched = []
    for inv in sorted(by_fund, key=str.lower):
        rs = by_fund[inv]
        dates = [r["data_date"] for r in rs if r["data_date"]]
        canoe_freqs = [r["frequency_canoe"] for r in rs if r["frequency_canoe"]]
        canoe_freq = max(set(canoe_freqs), key=canoe_freqs.count) if canoe_freqs else ""
        freq, source = infer_frequency(dates, canoe_freq)
        sponsors = [r["fund_sponsor"] for r in rs if r["fund_sponsor"]]
        sponsor = max(set(sponsors), key=sponsors.count) if sponsors else ""
        first = min(dates) if dates else None
        # Track from the start of the last full year of history, not inception --
        # the report stays readable and old one-off gaps don't show as overdue.
        start = max(first, date(date.today().year - 1, 1, 1)) if first else date.today()
        notes = f"auto-seeded ({source}; canoe says: {canoe_freq or 'n/a'})"
        last = max(dates) if dates else None
        if last and (date.today() - last).days > 200:
            notes += f"; last statement {last.isoformat()} -- verify still reporting"
        sched.append({
            "investment": inv,
            "fund_sponsor": sponsor,
            "frequency": freq,
            "grace_days": str(DEFAULT_GRACE.get(freq, 90)),
            "track": "yes" if freq != "none" else "no",
            "start_date": start.isoformat(),
            "doc_types": "",
            "notes": notes,
        })
    write_schedule(path, sched)
    return sched


SCHEDULE_HELP = [
    "Statement tracker schedule -- edit freely; the tracker re-reads this file each run.",
    "",
    "frequency : monthly | quarterly | annual | none",
    "track     : yes | no  (no = fund is ignored, e.g. wind-downs)",
    f"grace_days: days after period end before a missing statement is OVERDUE"
    f" (December period-ends get +{Q4_EXTRA_DAYS} automatically).",
    "start_date: first period the tracker should expect (YYYY-MM-DD).",
    "doc_types : optional ;-separated override of which document types satisfy"
    " a period for this fund (default: Account Statement, Capital Account"
    " Statement, Monthly/Quarterly/Annual Report, Financials).",
    "",
    "New funds appearing in Canoe are appended automatically with a NEW note.",
]


def write_schedule(path: str, sched: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.append(SCHEDULE_HEADER)
    for c in ws[1]:
        c.font = Font(bold=True)
    for s in sched:
        ws.append([s.get(k, "") for k in SCHEDULE_HEADER])
    # Dropdowns keep hand edits valid (rows beyond the current count included).
    last = max(len(sched) + 200, 500)
    dv_freq = DataValidation(type="list", formula1='"monthly,quarterly,annual,none"',
                             allow_blank=True)
    dv_track = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    ws.add_data_validation(dv_freq)
    ws.add_data_validation(dv_track)
    col_f = SCHEDULE_HEADER.index("frequency") + 1
    col_t = SCHEDULE_HEADER.index("track") + 1
    dv_freq.add(f"{openpyxl.utils.get_column_letter(col_f)}2:"
                f"{openpyxl.utils.get_column_letter(col_f)}{last}")
    dv_track.add(f"{openpyxl.utils.get_column_letter(col_t)}2:"
                 f"{openpyxl.utils.get_column_letter(col_t)}{last}")
    widths = {"A": 42, "B": 24, "C": 11, "D": 11, "E": 7, "F": 12, "G": 30, "H": 60}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    info = wb.create_sheet("How to use")
    for line in SCHEDULE_HELP:
        info.append([line])
    info.column_dimensions["A"].width = 110
    wb.save(path)


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def load_schedule(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Schedule"] if "Schedule" in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    header = [_cell_str(h) for h in next(rows, [])]
    out = []
    for r in rows:
        rec = {h: _cell_str(v) for h, v in zip(header, r) if h}
        if rec.get("investment"):
            out.append(rec)
    wb.close()
    return out


def load_schedule_csv(path: str) -> list[dict]:
    """Reader for the legacy csv schedule (pre-2026-08 layout); migration only."""
    with open(path, newline="") as f:
        reader = csv.DictReader(line for line in f if not line.startswith("#"))
        return [row for row in reader if (row.get("investment") or "").strip()]


def sync_new_funds(sched: list[dict], rows: list[dict], path: str) -> list[dict]:
    """Funds that appear in Canoe but not in the schedule get appended, flagged NEW."""
    known = {s["investment"] for s in sched}
    new_invs = sorted({r["investment"] for r in rows} - known, key=str.lower)
    if not new_invs:
        return sched
    seeded = seed_schedule(rows, path + ".tmp")
    os.remove(path + ".tmp")
    seeded_by_inv = {s["investment"]: s for s in seeded}
    for inv in new_invs:
        s = seeded_by_inv.get(inv)
        if s:
            s["notes"] = "NEW -- " + s["notes"]
            sched.append(s)
            print(f"  new fund added to schedule: {inv} ({s['frequency']})")
    write_schedule(path, sched)
    return sched


# --------------------------------------------------------------------------- #
# Reconciliation
# --------------------------------------------------------------------------- #

# Canoe's couldn't-identify buckets are not funds -- never track them. The
# documents in them need fixing in Canoe (see canoe_reclassify.py for the
# local-parse assist); until then they'd only add meaningless grid rows.
IGNORED_INVESTMENTS = {"unknown", "unknown investment"}

# Types that are entity-specific by nature: each investing entity gets its own.
# Only these can satisfy an ENTITY sub-row or trigger the amber re-tag flag --
# quarterly reports, financials etc. are addressed to all LPs, carry no entity,
# and count only at fund level.
ENTITY_DOC_TYPES = {"account statement", "capital account statement"}

# When several documents cover the same period, the one representing the actual
# statement wins (drives the grid's "Link" and the report detail): a capital
# account statement beats a quarterly report beats audited financials.
TYPE_PRIORITY = {
    "capital account statement": 0,
    "account statement": 1,
    "monthly report": 2,
    "quarterly report": 3,
    "annual report": 4,
    "financials": 5,
}

def _doc_rank(m: dict) -> tuple:
    return (TYPE_PRIORITY.get((m["document_type"] or "").lower(), 9),
            m["uploaded"] or date.min)

def reconcile(sched: list[dict], rows: list[dict], today: date) -> list[dict]:
    """Return one record per tracked fund x period with a status.

    Fund-level records (entity == "") count every statement for the fund,
    including ones Canoe never tagged to an entity. Funds where 2+ named
    entities invest additionally get per-entity records, each tracked from
    that entity's own first statement -- so each entity's statements can be
    checked (and linked) separately.
    """
    by_fund: dict[str, list[dict]] = {}
    for r in rows:
        by_fund.setdefault(r["investment"], []).append(r)

    out = []
    for s in sched:
        if (s.get("track") or "").strip().lower() not in ("yes", "y", "true", "1"):
            continue
        if s["investment"].strip().lower() in IGNORED_INVESTMENTS:
            continue
        freq = (s.get("frequency") or "").strip().lower()
        if freq not in ("monthly", "quarterly", "annual"):
            continue
        inv = s["investment"]
        try:
            grace = int(s.get("grace_days") or DEFAULT_GRACE[freq])
        except ValueError:
            grace = DEFAULT_GRACE[freq]
        start = parse_date(s.get("start_date")) or date(today.year - 1, 1, 1)

        fund_rows = by_fund.get(inv, [])
        named = sorted({r["entity"] for r in fund_rows
                        if r["entity"] and r["entity"] != "--"}, key=str.lower)
        groups = [("", fund_rows)]
        if len(named) >= 2:
            # Entity sub-rows only track entity-specific statement types;
            # fund-level reports can never turn an entity row green.
            groups += [(e, [r for r in fund_rows if r["entity"] == e
                            and r["document_type"].lower() in ENTITY_DOC_TYPES])
                       for e in named]

        # Account statements Canoe hasn't tagged to any entity, by period: a
        # red entity cell with one of these present becomes "retag" -- the
        # statement is in Canoe, it just needs its entity assigned. Untagged
        # fund-level docs (quarterly reports, financials) don't qualify --
        # they have no entity to assign.
        untagged_by_period: dict[str, list[dict]] = {}
        for r in fund_rows:
            if (r["entity"] in ("", "--") and r["data_date"]
                    and r["document_type"].lower() in ENTITY_DOC_TYPES):
                untagged_by_period.setdefault(
                    period_of(r["data_date"], freq), []).append(r)
        for docs_ in untagged_by_period.values():
            docs_.sort(key=_doc_rank)

        for entity, ent_rows in groups:
            ent_dates = [r["data_date"] for r in ent_rows if r["data_date"]]
            g_start = start
            if entity and ent_dates:
                g_start = max(start, min(ent_dates))

            docs_by_period: dict[str, list[dict]] = {}
            for r in ent_rows:
                if r["data_date"]:
                    docs_by_period.setdefault(period_of(r["data_date"], freq), []).append(r)

            for p in periods_between(g_start, today, freq):
                pe = period_end(p)
                due = pe + timedelta(days=grace + (Q4_EXTRA_DAYS if pe.month == 12 else 0))
                matched = sorted(docs_by_period.get(p, []), key=_doc_rank)
                clean = [m for m in matched
                         if m["document_status"].lower() not in REVIEW_STATUSES]
                flagged = [m for m in matched
                           if m["document_status"].lower() in REVIEW_STATUSES]
                if clean:
                    received = min((m["uploaded"] for m in clean if m["uploaded"]), default=None)
                    status = "received" if (received is None or received <= due) else "late"
                    best = clean[0]
                elif flagged:
                    status, received, best = "review", None, flagged[0]
                elif entity and untagged_by_period.get(p):
                    # Nothing tagged to this entity, but an untagged statement
                    # covers the period -- flag for re-tagging, don't cry missing.
                    status, received = "retag", None
                    best = untagged_by_period[p][0]
                else:
                    best, received = None, None
                    status = "overdue" if today > due else "pending"
                out.append({
                    "fund_sponsor": s.get("fund_sponsor") or "",
                    "investment": inv,
                    "entity": entity,
                    "frequency": freq,
                    "period": p,
                    "period_end": pe.isoformat(),
                    "due": due.isoformat(),
                    "status": status,
                    "received_date": received.isoformat() if received else "",
                    "data_date": best["data_date"].isoformat() if best and best["data_date"] else "",
                    "document_type": best["document_type"] if best else "",
                    "document_status": best["document_status"] if best else "",
                    "doc_name": best["doc_name"] if best else "",
                    "n_docs": len(matched),
                })
    return out


# --------------------------------------------------------------------------- #
# Outputs
# --------------------------------------------------------------------------- #

def write_status_csv(path: str, recs: list[dict]) -> None:
    if not recs:
        return
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(recs[0].keys()))
        w.writeheader()
        w.writerows(recs)


def write_received_log(path: str, rows: list[dict]) -> None:
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["investment", "fund_sponsor", "entity", "document_type",
                    "data_date", "uploaded", "document_status", "doc_name"])
        for r in sorted(rows, key=lambda r: (r["investment"].lower(),
                                             r["data_date"] or date.min)):
            w.writerow([r["investment"], r["fund_sponsor"], r["entity"],
                        r["document_type"],
                        r["data_date"].isoformat() if r["data_date"] else "",
                        r["uploaded"].isoformat() if r["uploaded"] else "",
                        r["document_status"], r["doc_name"]])


STATUS_META = {
    # status -> (label, pill background, pill text) -- muted, print-friendly
    "received": ("Received", "#dcefdc", "#1b5e20"),
    "late":     ("Received late", "#e3ecdf", "#33691e"),
    "pending":  ("Pending", "#fdf3d7", "#8a6d1a"),
    "overdue":  ("OVERDUE", "#fbe0dd", "#b71c1c"),
    "review":   ("Review", "#e8e0f5", "#4a148c"),
}


def write_html(path: str, recs: list[dict], undated: list[dict],
               periods_shown: int, generated: str) -> None:
    # The HTML detail stays at fund grain; entity detail lives in the grid
    # and the received log.
    recs = [r for r in recs if not r.get("entity")]
    freq_rank = {"monthly": 0, "quarterly": 1, "annual": 2}
    # column set = union of periods, most recent N, grouped by frequency
    by_freq: dict[str, list[dict]] = {}
    for r in recs:
        by_freq.setdefault(r["frequency"], []).append(r)

    counts = {k: 0 for k in STATUS_META}
    latest = {}   # per fund: most recent period record (drives the summary)
    for r in recs:
        key = (r["frequency"], r["investment"])
        if key not in latest or r["period_end"] > latest[key]["period_end"]:
            latest[key] = r
    for r in recs:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    open_overdue = sum(1 for r in recs if r["status"] == "overdue")
    review_n = sum(1 for r in recs if r["status"] == "review") + len(undated)
    pending_n = sum(1 for r in latest.values() if r["status"] == "pending")

    def esc(s):
        return html.escape(str(s))

    parts = [f"""<!-- generated by statement_tracker.py -->
<meta charset="utf-8">
<title>Statement Tracker</title>
<style>
 body {{ font: 14px/1.45 -apple-system, "Segoe UI", Helvetica, Arial, sans-serif;
        margin: 24px; color: #24292f; }}
 h1 {{ font-size: 20px; margin: 0 0 2px; }}
 .sub {{ color: #57606a; margin-bottom: 18px; }}
 .cards {{ display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 20px; }}
 .card {{ border: 1px solid #d0d7de; border-radius: 8px; padding: 10px 16px; min-width: 130px; }}
 .card b {{ display: block; font-size: 22px; }}
 table {{ border-collapse: collapse; margin-bottom: 28px; width: 100%; }}
 th, td {{ border: 1px solid #d8dee4; padding: 4px 8px; text-align: left; white-space: nowrap; }}
 th {{ background: #f6f8fa; position: sticky; top: 0; }}
 td.fund {{ max-width: 320px; overflow: hidden; text-overflow: ellipsis; }}
 .pill {{ display: inline-block; border-radius: 10px; padding: 1px 8px; font-size: 12px; }}
 .sponsor {{ color: #57606a; font-size: 12px; }}
 h2 {{ font-size: 16px; margin: 24px 0 8px; }}
 .overdue-row td {{ background: #fff5f4; }}
 caption {{ text-align: left; font-weight: 600; padding: 6px 0; }}
</style>
<h1>Statement Tracker</h1>
<div class="sub">Generated {esc(generated)} &middot; metadata-only (no document contents read) &middot;
edit <code>{SCHEDULE_FILE}</code> to change frequencies, grace periods, or tracked funds.</div>
<div class="cards">
 <div class="card"><b style="color:#b71c1c">{open_overdue}</b>overdue periods</div>
 <div class="card"><b style="color:#8a6d1a">{pending_n}</b>funds pending current period</div>
 <div class="card"><b style="color:#4a148c">{review_n}</b>needs review</div>
 <div class="card"><b style="color:#1b5e20">{counts.get('received',0)+counts.get('late',0)}</b>periods received</div>
</div>"""]

    # Exceptions table first -- the actionable list.
    exceptions = [r for r in recs if r["status"] in ("overdue", "review")]
    exceptions.sort(key=lambda r: (r["status"] != "overdue", r["due"]))
    if exceptions or undated:
        parts.append("<h2>Action needed</h2><table><tr><th>Status</th><th>Fund</th>"
                     "<th>Manager</th><th>Period</th><th>Due</th><th>Detail</th></tr>")
        for r in exceptions:
            lbl, bg, fg = STATUS_META[r["status"]]
            detail = (f'{esc(r["document_type"])} &middot; {esc(r["document_status"])}'
                      if r["status"] == "review" else
                      f'no statement since period end {esc(r["period_end"])}')
            parts.append(
                f'<tr class="overdue-row"><td><span class="pill" style="background:{bg};color:{fg}">{lbl}</span></td>'
                f'<td class="fund">{esc(r["investment"])}</td><td>{esc(r["fund_sponsor"])}</td>'
                f'<td>{esc(period_label(r["period"]))}</td><td>{esc(r["due"])}</td><td>{detail}</td></tr>')
        lbl, bg, fg = STATUS_META["review"]
        for r in undated:
            parts.append(
                f'<tr><td><span class="pill" style="background:{bg};color:{fg}">{lbl}</span></td>'
                f'<td class="fund">{esc(r["investment"])}</td><td>{esc(r["fund_sponsor"])}</td>'
                f'<td>&mdash;</td><td>&mdash;</td>'
                f'<td>undated statement &mdash; assign a data date in Canoe: {esc(r["doc_name"])}</td></tr>')
        parts.append("</table>")

    for freq in sorted(by_freq, key=lambda f: freq_rank.get(f, 9)):
        rs = by_freq[freq]
        periods = sorted({r["period"] for r in rs}, key=period_end)[-periods_shown:]
        funds = sorted({(r["fund_sponsor"], r["investment"]) for r in rs},
                       key=lambda t: (t[0].lower() or "~", t[1].lower()))
        cell = {(r["investment"], r["period"]): r for r in rs}
        parts.append(f'<h2>{freq.capitalize()} funds ({len(funds)})</h2><table>')
        parts.append("<tr><th>Fund</th>" +
                     "".join(f"<th>{esc(period_label(p))}</th>" for p in periods) + "</tr>")
        for sponsor, inv in funds:
            sp = f' <span class="sponsor">({esc(sponsor)})</span>' if sponsor and sponsor != inv else ""
            row = [f'<tr><td class="fund">{esc(inv)}{sp}</td>']
            for p in periods:
                r = cell.get((inv, p))
                if r is None:
                    row.append('<td style="color:#d0d7de">&mdash;</td>')
                else:
                    lbl, bg, fg = STATUS_META[r["status"]]
                    tip = esc(f'{r["doc_name"]} | data date {r["data_date"]} | uploaded {r["received_date"]}'
                              if r["doc_name"] else f'due {r["due"]}')
                    row.append(f'<td><span class="pill" title="{tip}" '
                               f'style="background:{bg};color:{fg}">{lbl}</span></td>')
            row.append("</tr>")
            parts.append("".join(row))
        parts.append("</table>")

    with open(path, "w") as f:
        f.write("\n".join(parts))


def _sanitize(component: str) -> str:
    # Mirrors canoe_bulk_download._sanitize so fund names map to archive folders.
    for ch in ("/", "\\", ":", "*", "?", '"', "<", ">", "|"):
        component = component.replace(ch, "-")
    return component.strip().strip(".") or "Unfiled"


def build_archive_index(dest: str) -> dict[str, str]:
    """Map lowercase file stem (dedup __N suffix stripped) -> archive relpath."""
    index: dict[str, str] = {}
    for root, dirs, files in os.walk(dest):
        dirs[:] = [d for d in dirs if not d.startswith(".") and d != SUBDIR]
        for f in files:
            if f.startswith(".") or f.startswith("~$"):
                continue
            base = re.sub(r"__\d+$", "", os.path.splitext(f)[0]).lower()
            index.setdefault(base, os.path.relpath(os.path.join(root, f), dest))
    return index


def _file_link(rec: dict, index: dict[str, str], dest: str) -> str | None:
    """Relative hyperlink (from the workbook's folder) to the statement file,
    falling back to the fund's archive folder."""
    rel = index.get((rec.get("doc_name") or "").lower())
    if rel is None:
        folder = _sanitize(rec["investment"])
        if os.path.isdir(os.path.join(dest, folder)):
            rel = folder
        else:
            return None
    return "../" + urllib.parse.quote(rel.replace(os.sep, "/"))


def write_xlsx(path: str, recs: list[dict], dest: str) -> None:
    """Simple received grid: one sheet per cadence, one row per fund, one column
    per period. Green = a statement for that period is in Canoe (click to open
    it), red = expected but not received, blank = not tracked for that period."""
    green = PatternFill("solid", fgColor="63BE7B")
    red = PatternFill("solid", fgColor="F8696B")
    amber = PatternFill("solid", fgColor="FFD966")
    thin = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    center = Alignment(horizontal="center")
    index = build_archive_index(dest)

    # NB: labels must not start with "=" or Excel treats them as formulas (#NAME?).
    LEGEND = [(green, 'Received -- click "Link" to open the statement'),
              (red, "Expected, not received in Canoe"),
              (amber, 'Statement in Canoe but not tagged to this entity -- click '
                      '"Tag" to open it, then assign the entity in Canoe'),
              (None, "Not tracked for this period")]
    NOTE = ("Funds with multiple investing entities show one sub-row per entity; "
            "the fund row is a header only -- status and links live on the entity "
            "rows. Amber cells mean the period's statement arrived but is not yet "
            "tagged to an entity in Canoe (it may belong to a different sub-row).")
    HDR_ROW = len(LEGEND) + 3          # legend, note, blank row, then the header

    def paint(c, rec):
        if rec["status"] == "retag":
            c.fill = amber
            link = _file_link(rec, index, dest)   # rec carries the untagged doc
            if link:
                c.value = "Tag"
                c.hyperlink = link
                c.font = Font(color="7F6000", underline="single")
                c.alignment = center
        elif int(rec["n_docs"] or 0) > 0:
            c.fill = green
            link = _file_link(rec, index, dest)
            if link:
                # Give the cell display text, otherwise Excel renders the raw URL.
                c.value = "Link"
                c.hyperlink = link
                c.font = Font(color="1B5E20", underline="single")
                c.alignment = center
        else:
            c.fill = red

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for freq, sheet_name in (("monthly", "Monthly"), ("quarterly", "Quarterly"),
                             ("annual", "Annual")):
        rs = [r for r in recs if r["frequency"] == freq]
        if not rs:
            continue
        ws = wb.create_sheet(sheet_name)
        periods = sorted({r["period"] for r in rs}, key=period_end)
        funds = sorted({r["investment"] for r in rs}, key=str.lower)
        cell_rec = {(r["investment"], r["entity"], r["period"]): r for r in rs}

        for i, (fill, label) in enumerate(LEGEND, start=1):
            sw = ws.cell(i, 2)
            sw.border = thin
            if fill:
                sw.fill = fill
            ws.cell(i, 3, label)
        ws.cell(len(LEGEND) + 1, 3, NOTE).font = Font(italic=True, color="666666")

        ws.cell(HDR_ROW, 1, "Fund / Entity").font = Font(bold=True)
        for j, p in enumerate(periods, start=2):
            c = ws.cell(HDR_ROW, j, period_label(p))
            c.font = Font(bold=True)
            c.alignment = center

        i = HDR_ROW + 1
        for inv in funds:
            entities = sorted({r["entity"] for r in rs
                               if r["investment"] == inv and r["entity"]}, key=str.lower)
            name = ws.cell(i, 1, inv)
            name.border = thin
            name.font = Font(bold=True)
            for j, p in enumerate(periods, start=2):
                c = ws.cell(i, j)
                c.border = thin
                # Funds with entity sub-rows: the roll-up row is a pure header;
                # all status color and links live on the entity rows.
                if entities:
                    continue
                rec = cell_rec.get((inv, "", p))
                if rec is not None:
                    paint(c, rec)
            i += 1
            for ent in entities:
                lbl = ws.cell(i, 1, "    " + ent)
                lbl.border = thin
                lbl.font = Font(color="444444")
                for j, p in enumerate(periods, start=2):
                    c = ws.cell(i, j)
                    c.border = thin
                    rec = cell_rec.get((inv, ent, p))
                    if rec is not None:
                        paint(c, rec)
                i += 1

        ws.column_dimensions["A"].width = 48
        for j in range(2, 2 + len(periods)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 10
        ws.freeze_panes = f"B{HDR_ROW + 1}"

    wb.save(path)


# --------------------------------------------------------------------------- #
# Digest: statements that arrived since the last run
# --------------------------------------------------------------------------- #

DIGEST_STATE = "digest_state.json"

def build_digest(backend: str, rows: list[dict], recs: list[dict]) -> tuple[str, list[dict]]:
    """Return (digest_html, new_rows). Tracks reported doc ids in a state file
    so each statement is announced exactly once, whatever the run cadence."""
    state_path = os.path.join(backend, DIGEST_STATE)
    seen: set = set()
    first_run = not os.path.exists(state_path)
    if not first_run:
        try:
            seen = set(json.load(open(state_path)).get("reported_ids", []))
        except (OSError, ValueError):
            first_run = True
    if first_run:
        # Baseline: don't announce years of history -- only the last 7 days.
        cutoff = date.today() - timedelta(days=7)
        new = [r for r in rows if r["uploaded"] and r["uploaded"] >= cutoff]
    else:
        new = [r for r in rows if r["doc_id"] and r["doc_id"] not in seen]
    new.sort(key=lambda r: (r["investment"].lower(), r["data_date"] or date.min))

    seen.update(r["doc_id"] for r in rows if r["doc_id"])
    json.dump({"updated": datetime.now(timezone.utc).isoformat(),
               "reported_ids": sorted(seen)}, open(state_path, "w"))

    n_over = sum(1 for r in recs if r["status"] == "overdue" and not r.get("entity"))
    esc = html.escape
    body = [f"""<meta charset="utf-8">
<style> body {{ font: 14px -apple-system, "Segoe UI", Helvetica, Arial, sans-serif; color: #24292f; }}
 table {{ border-collapse: collapse; }} th, td {{ border: 1px solid #d8dee4; padding: 4px 10px; text-align: left; }}
 th {{ background: #f6f8fa; }} </style>
<h2>Canoe statements digest &mdash; {esc(date.today().isoformat())}</h2>
<p><b>{len(new)}</b> new statement(s) since the last pull &middot; {n_over} period(s) currently overdue.
See the latest <i>Statement Tracker</i> workbook in the Canoe SharePoint library for the full picture.</p>"""]
    if new:
        body.append("<table><tr><th>Fund</th><th>Entity</th><th>Type</th>"
                    "<th>Period date</th><th>Uploaded</th></tr>")
        for r in new:
            ent = r["entity"] if r["entity"] not in ("", "--") else ""
            body.append(f"<tr><td>{esc(r['investment'])}</td><td>{esc(ent)}</td>"
                        f"<td>{esc(r['document_type'])}</td>"
                        f"<td>{esc(r['data_date'].isoformat() if r['data_date'] else 'undated')}</td>"
                        f"<td>{esc(r['uploaded'].isoformat() if r['uploaded'] else '')}</td></tr>")
        body.append("</table>")
    else:
        body.append("<p>No new statements this run.</p>")
    return "\n".join(body), new


def email_digest(digest_html: str, n_new: int) -> str:
    """Send the digest if SMTP settings are present in the environment/.env.
    Returns a short status string for the run log."""
    to = os.environ.get("CANOE_DIGEST_TO", "").strip()
    user = os.environ.get("CANOE_SMTP_USER", "").strip()
    password = os.environ.get("CANOE_SMTP_PASS", "").strip()
    if not (to and user and password):
        return "email not configured (set CANOE_DIGEST_TO / CANOE_SMTP_USER / CANOE_SMTP_PASS)"
    host = os.environ.get("CANOE_SMTP_HOST", "smtp.office365.com").strip()
    port = int(os.environ.get("CANOE_SMTP_PORT", "587"))
    sender = os.environ.get("CANOE_DIGEST_FROM", user).strip()

    msg = MIMEText(digest_html, "html")
    msg["Subject"] = f"Canoe statements digest -- {n_new} new ({date.today().isoformat()})"
    msg["From"] = sender
    msg["To"] = to
    try:
        with smtplib.SMTP(host, port, timeout=60) as s:
            s.starttls()
            s.login(user, password)
            s.sendmail(sender, [a.strip() for a in to.split(",") if a.strip()], msg.as_string())
        return f"emailed to {to}"
    except Exception as exc:                                  # noqa: BLE001
        return f"email FAILED ({exc.__class__.__name__}: {exc})"


# --------------------------------------------------------------------------- #
# Layout migration (pre-2026-08 file locations)
# --------------------------------------------------------------------------- #

def migrate_layout(dest: str, outdir: str, backend: str) -> None:
    """Move files from the old flat layout into backend/; one-time, idempotent."""
    moves = {
        os.path.join(outdir, CACHE_FILE): os.path.join(backend, CACHE_FILE),
        os.path.join(outdir, "Statement Tracker.html"): os.path.join(backend, "Statement Tracker.html"),
        os.path.join(outdir, "statement_status.csv"): os.path.join(backend, "statement_status.csv"),
        os.path.join(outdir, "statement_received_log.csv"): os.path.join(backend, "statement_received_log.csv"),
    }
    for src, dst in moves.items():
        if os.path.exists(src) and not os.path.exists(dst):
            os.replace(src, dst)
            print(f"  migrated    : {os.path.basename(src)} -> {BACKEND}/")
    old_csv = os.path.join(outdir, "statement_schedule.csv")
    new_sched = os.path.join(backend, SCHEDULE_FILE)
    if os.path.exists(old_csv):
        if not os.path.exists(new_sched):
            write_schedule(new_sched, load_schedule_csv(old_csv))
            print(f"  migrated    : statement_schedule.csv -> {BACKEND}/{SCHEDULE_FILE}")
        os.remove(old_csv)
    # The grid used to live at the archive root; it is regenerated in outdir now.
    old_grid = os.path.join(dest, GRID_PREFIX + ".xlsx")
    if os.path.exists(old_grid):
        os.remove(old_grid)
        print(f"  migrated    : removed old root copy of {GRID_PREFIX}.xlsx")


def archive_old_grids(outdir: str) -> None:
    """Sweep ALL previous grid workbooks into Archive/ so the folder always
    shows exactly one current workbook. Collision-safe."""
    archive_dir = os.path.join(outdir, ARCHIVE)
    for f in sorted(os.listdir(outdir)):
        if not (f.startswith(GRID_PREFIX) and f.endswith(".xlsx")):
            continue
        os.makedirs(archive_dir, exist_ok=True)
        target = os.path.join(archive_dir, f)
        root, ext = os.path.splitext(target)
        n = 2
        while os.path.exists(target):
            target = f"{root}__{n}{ext}"
            n += 1
        os.replace(os.path.join(outdir, f), target)
        print(f"  archived    : {f} -> {ARCHIVE}/{os.path.basename(target)}")


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main() -> None:
    ap = argparse.ArgumentParser(description="Track statements received per manager/fund.")
    ap.add_argument("--dest", required=True,
                    help="Archive root (the synced SharePoint 'Canoe' folder). "
                         f"Outputs go to <dest>/{SUBDIR}/.")
    ap.add_argument("--refresh", default="auto", choices=["auto", "full"],
                    help="auto = cached + delta by last-modified; full = re-pull everything.")
    ap.add_argument("--periods", type=int, default=13,
                    help="How many recent periods to show per grid (default 13).")
    args = ap.parse_args()

    dest = os.path.abspath(os.path.expanduser(args.dest))
    outdir = os.path.join(dest, SUBDIR)
    backend = os.path.join(outdir, BACKEND)
    os.makedirs(backend, exist_ok=True)
    today = date.today()

    print("Statement tracker")
    print(f"  grid        : {os.path.join(outdir, GRID_PREFIX)} <date>.xlsx")
    print(f"  backend     : {backend}")
    migrate_layout(dest, outdir, backend)

    # The schedule is read first: per-fund doc_types overrides extend which
    # document types the metadata pull must cover.
    sched_path = os.path.join(backend, SCHEDULE_FILE)
    sched = load_schedule(sched_path) if os.path.exists(sched_path) else None
    overrides: dict = {}
    type_names = list(DEFAULT_STATEMENT_TYPE_NAMES)
    known = {t.lower() for t in type_names}
    for s in sched or []:
        types = [t.strip() for t in (s.get("doc_types") or "").split(";") if t.strip()]
        if types:
            overrides[s["investment"]] = {t.lower() for t in types}
            for t in types:
                if t.lower() not in known:
                    known.add(t.lower())
                    type_names.append(t)

    docs = load_metadata(os.path.join(backend, CACHE_FILE), args.refresh, type_names)
    print(f"  documents   : {len(docs)} across {len(type_names)} statement types (all categories)")

    routed = merrill_stems(dest)
    base_rows = statement_rows(docs, {}, routed)
    if sched is None:
        print("  no schedule found -- seeding from history "
              f"(review and edit {SCHEDULE_FILE}!)")
        sched = seed_schedule(base_rows, sched_path)
    else:
        sched = sync_new_funds(sched, base_rows, sched_path)
    rows = statement_rows(docs, overrides, routed) if overrides else base_rows
    print(f"  statements  : {len(rows)} fund-allocation rows "
          f"across {len({r['investment'] for r in rows})} funds")

    recs = reconcile(sched, rows, today)
    tracked = {s["investment"] for s in sched
               if (s.get("track") or "").strip().lower() in ("yes", "y", "true", "1")}
    undated = sorted((r for r in rows if r["data_date"] is None and r["investment"] in tracked),
                     key=lambda r: r["investment"].lower())
    fund_recs = [r for r in recs if not r["entity"]]
    n_over = sum(1 for r in fund_recs if r["status"] == "overdue")
    n_review = sum(1 for r in fund_recs if r["status"] == "review")
    print(f"  reconciled  : {len(fund_recs)} fund-periods "
          f"({len(recs) - len(fund_recs)} entity sub-rows) | overdue {n_over} | "
          f"review {n_review} | undated {len(undated)}")

    generated = datetime.now().strftime("%Y-%m-%d %H:%M")
    write_status_csv(os.path.join(backend, "statement_status.csv"), recs)
    write_received_log(os.path.join(backend, "statement_received_log.csv"), rows)
    write_html(os.path.join(backend, "Statement Tracker.html"), recs, undated, args.periods, generated)
    # The grid is the team's view -- EVERY run writes a brand-new workbook and
    # archives the previous one, same-day reruns included. A reused filename is
    # a reused OneDrive item, and rewriting an item someone has open in Excel
    # wedges its sync; a fresh name always uploads.
    archive_old_grids(outdir)
    base = f"{GRID_PREFIX} {today.isoformat()}"
    archive_dir = os.path.join(outdir, ARCHIVE)
    prior_today = sum(1 for f in os.listdir(archive_dir) if f.startswith(base)) \
        if os.path.isdir(archive_dir) else 0
    grid_name = f"{base}.xlsx" if prior_today == 0 else f"{base} ({prior_today + 1}).xlsx"
    write_xlsx(os.path.join(outdir, grid_name), recs, dest)
    print(f"  wrote       : {grid_name} + backend detail (html, csvs)")

    digest_html, new_rows = build_digest(backend, rows, recs)
    with open(os.path.join(backend, "Statement Digest.html"), "w") as f:
        f.write(digest_html)
    print(f"  digest      : {len(new_rows)} new statement(s); {email_digest(digest_html, len(new_rows))}")


if __name__ == "__main__":
    main()
